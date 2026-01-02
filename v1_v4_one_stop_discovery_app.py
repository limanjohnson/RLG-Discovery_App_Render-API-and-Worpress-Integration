# v3_one_stop_discovery_app.py
# Run with: python -m streamlit run v3_one_stop_discovery_app.py

from __future__ import annotations

import io, os, re, csv, zipfile, tempfile, shutil, json, platform
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Optional, Tuple, List, Iterable, Set
from collections import Counter

import streamlit as st
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF / imaging
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.colors import Color
from PIL import Image, ImageDraw, ImageFont, ImageOps

# Optional: pdf2image (for previews & OCR fallback)
try:
    from pdf2image import convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except Exception:
    PDF2IMAGE_AVAILABLE = False

# Optional: pikepdf for unlocking & repair
try:
    import pikepdf
    from pikepdf import PasswordError, PdfError
    PIKEPDF_AVAILABLE = True
except Exception:
    pikepdf = None
    PasswordError = PdfError = Exception
    PIKEPDF_AVAILABLE = False

# Optional: PyMuPDF
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

# Optional: OCR
try:
    import pytesseract
except Exception:
    pytesseract = None

def load_font(font_name: str, size: int) -> ImageFont.FreeTypeFont:
    """
    Load a font from common system paths or fallback to default.
    """
    # Common paths for Arial or similar sans-serif
    candidates = [
        # User requested specific path (if provided as absolute path)
        font_name,
        # macOS
        f"/Library/Fonts/{font_name}.ttf",
        f"/System/Library/Fonts/{font_name}.ttf",
        # Linux (Debian/Ubuntu/Render)
        f"/usr/share/fonts/truetype/msttcorefonts/{font_name}.ttf",
        f"/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", # Fallback 1
        f"/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", # Fallback 2
    ]
    
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
            
    # If all else fails, try loading by name (OS might resolve it)
    try:
        return ImageFont.truetype(f"{font_name}.ttf", size)
    except Exception:
        pass
        
    return ImageFont.load_default()

# ------------------------
# App config
# ------------------------
st.set_page_config(page_title="Discovery One-Stop", layout="wide")

# ------------------------
# Global helpers & project memory
# ------------------------
def _download_bytes(label: str, data: bytes, filename: str, mime: str):
    st.download_button(label, data=data, file_name=filename, mime=mime, use_container_width=True)


def _zip_dir(dir_path: Path) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in dir_path.rglob("*"):
            if p.is_file():
                zf.write(p, p.relative_to(dir_path))
    buf.seek(0)
    return buf.read()


def _zip_from_pairs(pairs: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rel, b in pairs:
            if rel.endswith("/"):
                continue
            zf.writestr(rel, b)
    buf.seek(0)
    return buf.read()


def _color_from_hex(hex_str: str) -> Tuple[int, int, int]:
    hex_str = hex_str.strip("#")
    if len(hex_str) == 6:
        r = int(hex_str[0:2], 16)
        g = int(hex_str[2:4], 16)
        b = int(hex_str[4:6], 16)
        return (r, g, b)
    return (0, 0, 255)


def natural_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]


def _detect_poppler_path() -> Optional[str]:
    for c in ("/opt/homebrew/bin", "/usr/local/bin"):
        if Path(c, "pdftoppm").exists():
            return c
    return None


def _format_label(prefix: str, number: int, digits: int, with_space: bool = True) -> str:
    return f"{prefix}{' ' if with_space else ''}{number:0{digits}d}"


def _pil_dpi(image: Image.Image) -> float:
    dpi = image.info.get("dpi")
    if isinstance(dpi, tuple) and dpi and dpi[0]:
        return float(dpi[0])
    return 150.0


def _preview_image_from_pdf_bytes(pdf_bytes: bytes, page_index: int, poppler_path: Optional[str] = None) -> Optional[Image.Image]:
    if not PDF2IMAGE_AVAILABLE:
        return None
    try:
        kwargs = dict(first_page=page_index + 1, last_page=page_index + 1, dpi=150)
        if poppler_path:
            kwargs["poppler_path"] = poppler_path
        imgs = convert_from_bytes(pdf_bytes, **kwargs)
        if imgs:
            return imgs[0]
    except Exception:
        return None
    return None

# ---------- macOS junk filters ----------
def _is_mac_resource_junk(path_str: str) -> bool:
    p = str(path_str).replace("\\", "/")
    base = Path(p).name
    if "/__MACOSX/" in p:
        return True
    if base.startswith("._"):
        return True
    if base.lower() == ".ds_store":
        return True
    return False


def _filter_pairs_nonjunk(pairs: List[Tuple[str, bytes]]) -> List[Tuple[str, bytes]]:
    return [(rel, b) for rel, b in pairs if not _is_mac_resource_junk(rel)]


# ---------- Project memory ----------
def project_init():
    ss = st.session_state
    ss.setdefault("project_files", None)          # List[Tuple[str, bytes]]
    ss.setdefault("project_origin", "")           # description string
    ss.setdefault("bates_records_df", None)       # DataFrame
    ss.setdefault("labeled_files", None)          # List[Tuple[str, bytes]]
    ss.setdefault("labeled_zip", None)            # bytes
    ss.setdefault("last_action", "")              # status text


def project_set(files: List[Tuple[str, bytes]], origin: str):
    clean = _filter_pairs_nonjunk(files)
    st.session_state["project_files"] = clean
    st.session_state["project_origin"] = origin
    st.session_state["last_action"] = f"Project updated from {origin}"


def project_clear():
    for k in ["project_files", "project_origin", "bates_records_df", "labeled_files", "labeled_zip", "last_action"]:
        st.session_state[k] = None if k != "project_origin" else ""


def project_summary() -> str:
    ss = st.session_state
    n = len(ss["project_files"]) if ss.get("project_files") else 0
    base = f"{n} file(s) in current project"
    origin = ss.get("project_origin") or "—"
    return f"{base} • Source: {origin}"

# --- Date Produced helper: parse from 2nd-level subfolder under the parent ---
def _parse_date_from_text(text: str) -> Optional[date]:
    if not text:
        return None
    # yyyy.mm.dd | yyyy-mm-dd | yyyy_mm_dd
    for m in re.finditer(r"\b(20\d{2}|19\d{2})[._/-](\d{1,2})[._/-](\d{1,2})\b", text):
        y, mo, d = map(int, m.groups())
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            pass
    # yyyymmdd
    m = re.search(r"\b(20\d{2}|19\d{2})(\d{2})(\d{2})\b", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            pass
    # mm.dd.yyyy | etc
    for m in re.finditer(r"\b(\d{1,2})[._/-](\d{1,2})[._/-](20\d{2}|19\d{2})\b", text):
        mo, d, y = map(int, m.groups())
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            pass
    return None


def _extract_date_produced_from_rel(rel_dir: str, filename: str = "") -> Optional[date]:
    parts = Path(rel_dir).parts if rel_dir else ()
    candidates: List[str] = []
    # Explicit: 2nd-level inside the parent folder (index 1)
    if len(parts) >= 2:
        candidates.append(parts[1])
    # fallbacks
    candidates.extend(parts)
    if filename:
        candidates.append(filename)
    for token in candidates:
        d = _parse_date_from_text(token)
        if d:
            return d
    return None

# ---------------- Smarter Bates detection (avoid random numbers) ----------------
_BLACKLIST_PREFIXES = {
    "MONTHLY", "BOX", "ID", "TARGET", "REQUESTED", "MISC"
}

_CANDIDATE_BATES_RE = re.compile(
    r"\b([A-Z][A-Z0-9. ]{1,30}?)[\s\-–—]*([0-9]{6,10})\b"
)

def _normalize_prefix(s: str) -> str:
    s = re.sub(r"\s+", " ", s).strip(" -–—.")
    return s.upper()

def _is_zero_padded(num: str) -> bool:
    return len(num) >= 6 and num[0] == "0"

def _extract_candidates(text: str) -> List[Tuple[str, str]]:
    out: List[Tuple[str,str]] = []
    if not text:
        return out
    for m in _CANDIDATE_BATES_RE.finditer(text.upper()):
        pfx = _normalize_prefix(m.group(1))
        num = m.group(2)
        out.append((pfx, num))
    return out

def _choose_dominant_prefix(cands: List[Tuple[str, str]]) -> Optional[str]:
    if not cands:
        return None
    zp = [p for p,n in cands if _is_zero_padded(n) and p not in _BLACKLIST_PREFIXES]
    if zp:
        return Counter(zp).most_common(1)[0][0]
    nb = [p for p,_ in cands if p not in _BLACKLIST_PREFIXES]
    if nb:
        return Counter(nb).most_common(1)[0][0]
    return None

def _best_token_for_prefix(cands: List[Tuple[str,str]], want_prefix: str) -> Optional[str]:
    for pfx, num in cands:
        if pfx == want_prefix and _is_zero_padded(num):
            return f"{pfx} {num}"
    for pfx, num in cands:
        if pfx == want_prefix:
            return f"{pfx} {num}"
    return None

def _pdf_page_text_or_ocr(pdf_bytes: bytes, page_index_zero: int) -> str:
    if fitz is not None:
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            if 0 <= page_index_zero < doc.page_count:
                page = doc.load_page(page_index_zero)
                txt = page.get_text("text") or ""
                if txt.strip():
                    doc.close()
                    return txt
                if pytesseract is not None:
                    pix = page.get_pixmap()
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    ocr_txt = pytesseract.image_to_string(img) or ""
                    doc.close()
                    return ocr_txt
            doc.close()
        except Exception:
            pass
    if PDF2IMAGE_AVAILABLE and pytesseract is not None:
        try:
            imgs = convert_from_bytes(pdf_bytes, first_page=page_index_zero+1, last_page=page_index_zero+1, dpi=200)
            if imgs:
                return pytesseract.image_to_string(imgs[0]) or ""
        except Exception:
            pass
    return ""

def _image_bytes_text_ocr(img_bytes: bytes) -> str:
    try:
        with Image.open(io.BytesIO(img_bytes)) as im:
            im = ImageOps.exif_transpose(im)
            if pytesseract is not None:
                return pytesseract.image_to_string(im) or ""
    except Exception:
        pass
    return ""

def _extract_bates_for_file(rel_path: str, data: bytes) -> Tuple[Optional[str], Optional[str]]:
    ext = Path(rel_path).suffix.lower()

    if ext == ".pdf":
        txt1 = _pdf_page_text_or_ocr(data, 0)
        c1 = _extract_candidates(txt1)

        page_count = None
        if fitz is not None:
            try:
                d = fitz.open(stream=data, filetype="pdf"); page_count = d.page_count; d.close()
            except Exception:
                pass
        if page_count is None:
            try:
                page_count = len(PdfReader(io.BytesIO(data)).pages)
            except Exception:
                page_count = 1
        last_index = max(0, page_count - 1)

        txtN = _pdf_page_text_or_ocr(data, last_index)
        cN = _extract_candidates(txtN)

        all_cands = c1 + cN
        dom = _choose_dominant_prefix(all_cands)
        if not dom:
            return None, None

        first_tok = _best_token_for_prefix(c1, dom) or _best_token_for_prefix(all_cands, dom)
        last_tok  = _best_token_for_prefix(cN, dom) or _best_token_for_prefix(all_cands[::-1], dom)

        if not first_tok and not last_tok:
            return None, None
        if first_tok and not last_tok:
            return first_tok, first_tok
        if last_tok and not first_tok:
            return last_tok, last_tok

        def _num_part(tok: str) -> int:
            return int(re.search(r"(\d{6,10})$", tok).group(1))

        try:
            n1 = _num_part(first_tok); n2 = _num_part(last_tok)
            if n2 < n1:
                first_tok, last_tok = last_tok, first_tok
        except Exception:
            pass
        return first_tok, last_tok

    elif ext in {".jpg", ".jpeg", ".png"}:
        txt = _image_bytes_text_ocr(data)
        cands = _extract_candidates(txt)
        dom = _choose_dominant_prefix(cands)
        if dom:
            tok = _best_token_for_prefix(cands, dom)
            return tok, tok
        return None, None

    else:
        return None, None

def _scan_pairs_for_bates(pairs: List[Tuple[str, bytes]]) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    progress = st.progress(0.0, text="Scanning files for Bates labels…")
    total = max(1, len(pairs))
    for i, (rel, b) in enumerate(pairs, start=1):
        if _is_mac_resource_junk(rel):
            continue
        p = Path(rel)
        rel_dir = str(p.parent) if str(p.parent) != "." else ""
        fname = p.name
        try:
            first, last = _extract_bates_for_file(rel, b)
        except Exception:
            first, last = None, None
        rows.append({
            "rel_dir": rel_dir,
            "filename": fname,
            "first_label": first or "",
            "last_label": last or ""
        })
        progress.progress(min(i/total, 1.0))
    progress.empty()
    return pd.DataFrame(rows)

# ======================================================
# 1) Unlock PDFs — with junk filters
# ======================================================
def ui_unlocker():
    st.subheader("Unlock PDFs")
    st.caption("Remove encryption from PDFs you are authorized to access.")

    if not PIKEPDF_AVAILABLE:
        st.error("pikepdf is not installed in this environment. Install it first:\n\npip install pikepdf")
        return

    st.markdown("#### Source")
    files = st.file_uploader(
        "Drag & drop one or more PDF files or .zip archives (nested folders OK)",
        type=["pdf", "zip"],
        accept_multiple_files=True,
        help="Drop parent folders as .zip to process everything recursively.",
        key="unlock_files",
    )

    st.markdown("#### Password Mode")
    password_mode = st.radio(
        "How will you provide passwords?",
        ["Single password for all", "Per-file password list (CSV)", "Try no password (for unencrypted files)"],
        horizontal=False,
        key="unlock_mode",
    )

    password_for_all: Optional[str] = None
    password_map: Dict[str, str] = {}

    if password_mode == "Single password for all":
        password_for_all = st.text_input(
            "Password (applied to every file)",
            type="password",
            help="Leave blank only if your PDFs are actually not encrypted.",
            key="unlock_pw_all",
        )
    elif password_mode == "Per-file password list (CSV)":
        st.markdown(
            "Upload a CSV with two columns: **filename,password**. "
            "Filenames can match the full path inside the ZIP, the basename, or the stem."
        )
        csv_file = st.file_uploader("Upload CSV", type=["csv"], accept_multiple_files=False, key="unlock_pw_csv")
        if csv_file is not None:
            import csv as _csv
            decoded = csv_file.read().decode("utf-8", errors="replace")
            reader = _csv.reader(decoded.splitlines())
            header = next(reader, None)
            if header and len(header) >= 2:
                try:
                    fn_idx = header.index("filename")
                    pw_idx = header.index("password")
                except ValueError:
                    fn_idx, pw_idx = 0, 1
                for row in reader:
                    if len(row) >= 2:
                        password_map[row[fn_idx]] = row[pw_idx]
            else:
                for row in reader:
                    if len(row) >= 2:
                        password_map[row[0]] = row[1]
    else:
        st.info("The app will attempt to open each PDF without a password. Encrypted PDFs will fail gracefully.")

    st.divider()
    start = st.button("Unlock PDFs and build ZIP", type="primary", disabled=not files, key="unlock_start")

    def _resolve_password(path: str) -> Optional[str]:
        if password_mode == "Single password for all":
            return (password_for_all or "").strip() or None
        elif password_mode == "Per-file password list (CSV)":
            base = os.path.basename(path)
            stem = os.path.splitext(base)[0]
            return password_map.get(path) or password_map.get(base) or password_map.get(stem)
        else:
            return None

    def _process_pdf(src_bytes: bytes, password: Optional[str]) -> Tuple[str, Optional[bytes]]:
        try:
            with io.BytesIO(src_bytes) as src_buf:
                try:
                    pdf = pikepdf.open(src_buf) if password is None else pikepdf.open(src_buf, password=password)
                except PasswordError:
                    return "❌ Password required or incorrect", None
                except PdfError as e:
                    return f"❌ PDF error: {e.__class__.__name__}", None
                out_mem = io.BytesIO()
                pdf.save(out_mem)  # saved without encryption
                pdf.close()
                return "✅ Unlocked", out_mem.getvalue()
        except Exception as e:
            return f"❌ Unexpected error: {e}", None

    if start:
        result_rows = []
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for f in files or []:
                fname = f.name
                if _is_mac_resource_junk(fname):
                    result_rows.append({"file": fname, "output": "—", "status": "⏭️ Skipped macOS resource file"})
                    continue
                if fname.lower().endswith(".pdf"):
                    status, data = _process_pdf(f.read(), _resolve_password(fname))
                    out_name = os.path.splitext(fname)[0] + "_unlocked.pdf"
                    if data is not None:
                        zf.writestr(out_name, data)
                    result_rows.append({"file": fname, "output": out_name if data else "—", "status": status})
                elif fname.lower().endswith(".zip"):
                    try:
                        with zipfile.ZipFile(io.BytesIO(f.read()), 'r') as inzip:
                            for member in inzip.namelist():
                                if member.endswith('/'):
                                    continue
                                if _is_mac_resource_junk(member):
                                    continue
                                if not member.lower().endswith('.pdf'):
                                    continue
                                pw = _resolve_password(member)
                                status, data = _process_pdf(inzip.read(member), pw)
                                out_name = f"{os.path.splitext(member)[0]}_unlocked.pdf"
                                if data is not None:
                                    zf.writestr(out_name, data)
                                result_rows.append({"file": member, "output": out_name if data else "—", "status": status})
                    except zipfile.BadZipFile:
                        result_rows.append({"file": fname, "output": "—", "status": "❌ Not a valid ZIP"})
                else:
                    result_rows.append({"file": fname, "output": "—", "status": "❌ Unsupported type"})

        any_ok = any(r["status"].startswith("✅") for r in result_rows)
        if any_ok:
            zip_buffer.seek(0)
            _download_bytes(
                "Download unlocked PDFs (ZIP)",
                data=zip_buffer.getvalue(),
                filename="unlocked_pdfs.zip",
                mime="application/zip",
            )

        if result_rows:
            st.markdown("#### Results")
            st.dataframe(pd.DataFrame(result_rows), use_container_width=True)
        else:
            st.info("No results to show yet.")

# ======================================================
# 2) Organize by Year
# ======================================================
MONTHS = r"Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?"
PATTERNS: List[re.Pattern] = [
    re.compile(rf"(?<!\d)((?P<year>20\d{{2}}|19\d{{2}}))(?!\d)", re.IGNORECASE),
    re.compile(rf"(?<!\d)\d{{1,2}}[ _\.-]\d{{1,2}}[ _\.-](?P<year>20\d{{2}}|19\d{{2}})(?!\d)", re.IGNORECASE),
    re.compile(rf"(?:(?:{MONTHS})\s*\d{{1,2}}[ ,._-]*)?(?P<year>20\d{{2}}|19\d{{2}})", re.IGNORECASE),
]

def preprocess_filename(name: str) -> str:
    return re.sub(r"^[A-Za-z]*\d{4,}[ _\.-]*", "", name)

def extract_year_from_name(name: str, min_year: int, max_year: int, year_policy: str = "first") -> Tuple[Optional[int], str]:
    name = preprocess_filename(name)
    candidates: List[Tuple[int, str, Tuple[int, int]]] = []
    for idx, pat in enumerate(PATTERNS, start=1):
        for m in pat.finditer(name):
            y = int(m.group("year"))
            if min_year <= y <= max_year:
                candidates.append((y, f"pattern{idx}", m.span()))
    if not candidates:
        return None, "no-year-found"
    if year_policy == "max":
        chosen = max(candidates, key=lambda t: t[0])
    elif year_policy == "last":
        chosen = candidates[-1]
    else:
        chosen = candidates[0]
    year, patname, span = chosen
    return year, f"{patname}@{span}"

def ui_organizer():
    st.subheader("Organize by Year")

    left, right = st.columns([0.95, 1.45], gap="large")

    with left:
        st.markdown("#### Source")
        source = st.radio("Choose source", ["Use current project", "Upload"], horizontal=True, key="org_source")
        uploaded: List[Tuple[str, bytes]] = []

        if source == "Upload":
            mode = st.radio("Input", ["Multiple PDFs", "ZIP of a folder"], horizontal=True, key="org_mode")
            if mode == "Multiple PDFs":
                files = st.file_uploader("Upload PDFs", type=["pdf"], accept_multiple_files=True, key="org_files")
                if files:
                    for f in files:
                        if _is_mac_resource_junk(f.name):
                            continue
                        uploaded.append((f.name, f.read()))
            else:
                z = st.file_uploader("Upload ZIP", type=["zip"], key="org_zip")
                if z:
                    with zipfile.ZipFile(io.BytesIO(z.read())) as zf:
                        for info in zf.infolist():
                            if info.is_dir():
                                continue
                            if _is_mac_resource_junk(info.filename):
                                continue
                            if not info.filename.lower().endswith(".pdf"):
                                continue
                            uploaded.append((info.filename, zf.read(info)))
        else:
            if st.session_state.get("project_files"):
                uploaded = _filter_pairs_nonjunk(st.session_state["project_files"])
            else:
                st.info("No project files yet. Upload in Unlock or set from another tab.")
                uploaded = []

        st.divider()
        st.markdown("#### Rules")
        min_year = st.number_input("Min year", 1900, 3000, 1900)
        max_year = st.number_input("Max year", 1900, 3000, 2099)
        year_policy = st.selectbox("When multiple years appear", ["first", "last", "max"], index=0, key="org_policy")
        unknown_folder = st.text_input("Folder for files without a year", "Unknown", key="org_unknown")

        st.divider()
        colA, colB = st.columns(2)
        with colA:
            run = st.button("Organize", type="primary", use_container_width=True, disabled=not uploaded, key="org_run")
        with colB:
            adopt = st.checkbox("Set result as current project", value=True, key="org_adopt")

    with right:
        st.markdown("#### Result")
        if run:
            with tempfile.TemporaryDirectory() as tmp_dir:
                tmp = Path(tmp_dir)
                out_root = tmp / f"organized_{datetime.now().strftime('%Y%m%d-%H%M%S')}"
                out_root.mkdir(parents=True, exist_ok=True)
                report_rows = []
                pairs: List[Tuple[str, bytes]] = []

                for display_name, data in uploaded:
                    year, reason = extract_year_from_name(Path(display_name).name, min_year, max_year, year_policy)
                    folder = str(year) if year is not None else unknown_folder
                    target_dir = out_root / folder
                    target_dir.mkdir(parents=True, exist_ok=True)
                    dest = target_dir / Path(display_name).name
                    i = 1
                    while dest.exists():
                        dest = target_dir / f"{Path(display_name).stem}__{i}{Path(display_name).suffix}"
                        i += 1
                    dest.write_bytes(data)
                    pairs.append((str(dest.relative_to(out_root)), data))
                    report_rows.append({
                        "source_name": display_name,
                        "dest_folder": folder,
                        "final_dest": str(dest.relative_to(out_root)),
                        "year": year or "",
                        "reason": reason
                    })

                zbytes = _zip_dir(out_root)
                st.success("Done")
                _download_bytes("Download organized ZIP", zbytes, "organized_by_year.zip", "application/zip")
                st.dataframe(pd.DataFrame(report_rows), use_container_width=True)

                if adopt and pairs:
                    project_set(pairs, origin="Organize by Year")
                    st.success("Set as current project.")
        else:
            if st.session_state.get("project_files"):
                st.info(project_summary())
            else:
                st.info("Upload or use current project, set rules, then click Organize.")

# ======================================================
# 3) Bates Labeler (Zones + Safety Border) with project memory
# ======================================================
IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

@dataclass
class BatesRecord:
    rel_dir: str
    filename: str
    pages_or_files: int
    first_label: str
    last_label: str
    category: str  # deepest folder

def _page_size(page) -> Tuple[float, float]:
    return float(page.mediabox.width), float(page.mediabox.height)

def _overlay_pdf(
    label: str,
    w: float, h: float,
    font_name: str, font_size: int,
    margin_right: float, margin_bottom: float,
    color_rgb: Tuple[int,int,int],
    left_punch_margin: float = 0.0,
    border_all_pt: float = 0.0,
) -> PdfReader:
    from io import BytesIO
    r, g, b = color_rgb
    packet = BytesIO()
    can = rl_canvas.Canvas(packet, pagesize=(w, h))

    if border_all_pt and border_all_pt > 0:
        can.setFillColor(Color(1, 1, 1))
        B = float(border_all_pt)
        can.rect(0, h - B, w, B, stroke=0, fill=1)
        can.rect(0, 0, w, B, stroke=0, fill=1)
        can.rect(0, 0, B, h, stroke=0, fill=1)
        can.rect(w - B, 0, B, h, stroke=0, fill=1)

    if left_punch_margin and left_punch_margin > 0:
        can.setFillColor(Color(1, 1, 1))
        can.rect(0, 0, left_punch_margin, h, stroke=0, fill=1)

    can.setFont(font_name, font_size)
    can.setFillColor(Color(r/255, g/255, b/255))
    eff_mr = max(margin_right, border_all_pt or 0.0)
    eff_mb = max(margin_bottom, border_all_pt or 0.0)
    can.drawRightString(w - eff_mr, eff_mb, label)

    can.save()
    packet.seek(0)
    return PdfReader(packet)

def _label_image(
    in_file: Path, out_file: Path, label: str,
    font_name: str, font_size_pt: int,
    margin_right_pt: float, margin_bottom_pt: float,
    color_rgb: Tuple[int,int,int],
    left_punch_margin_pt: float = 0.0,
    border_all_pt: float = 0.0,
):
    """Apply a Bates label to an image file.

    Font size is based on the requested point size *and* on image dimensions,
    so labels stay readable even on very large photos.
    """
    img = Image.open(in_file)
    img = ImageOps.exif_transpose(img)

    # Base conversion from points to pixels using DPI
    dpi = _pil_dpi(img)
    px_per_point = dpi / 72.0

    mx = int(round(margin_right_pt * px_per_point))
    my = int(round(margin_bottom_pt * px_per_point))
    lp = int(round(left_punch_margin_pt * px_per_point))
    bp = int(round(border_all_pt * px_per_point))

    # Optional left margin for 3-hole punch
    if lp > 0:
        new_img = Image.new(
            "RGB" if img.mode == "RGB" else "RGBA",
            (img.width + lp, img.height),
            (255, 255, 255) if img.mode != "RGBA" else (255, 255, 255, 0),
        )
        new_img.paste(img, (lp, 0))
        img = new_img

    # Optional all-sides white border
    if bp > 0:
        new_img = Image.new(
            "RGB" if img.mode == "RGB" else "RGBA",
            (img.width + 2 * bp, img.height + 2 * bp),
            (255, 255, 255) if img.mode != "RGBA" else (255, 255, 255, 0),
        )
        new_img.paste(img, (bp, bp))
        img = new_img
        mx = max(mx, bp)
        my = max(my, bp)

    # --- Font size: ensure it's big enough for large images ---
    # 1) What the user asked for, in pixels (via DPI)
    fs_from_points = font_size_pt * px_per_point
    # 2) A minimum relative size: ~2.5% of the smaller image dimension
    relative_min = 0.025 * min(img.width, img.height)
    # Final pixel size is the max of the two, with a floor of 10px
    fs_px = int(max(10, round(max(fs_from_points, relative_min))))

    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGBA")

    draw = ImageDraw.Draw(img)
    try:
        font = load_font(font_name, fs_px)
    except Exception:
        font = ImageFont.load_default()

    # Measure text to position within margins
    bbox = draw.textbbox((0, 0), label, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    x = max(0, img.width - mx - tw)
    y = max(0, img.height - my - th)

    # Outline for contrast
    for ox, oy in ((-1, 0), (1, 0), (0, -1), (0, 1)):
        draw.text((x + ox, y + oy), label, font=font, fill=(0, 0, 0))

    # Main text
    draw.text((x, y), label, font=font, fill=color_rgb)

    # Save out with same extension
    if out_file.suffix.lower() in [".jpg", ".jpeg"]:
        img.convert("RGB").save(out_file, quality=92, optimize=True)
    else:
        img.save(out_file)

def walk_and_label(
    input_zip_or_pdfs: List[Tuple[str, bytes]], *,
    prefix: str, start_num: int, digits: int,
    font_name: str, font_size: int,
    margin_right: float, margin_bottom: float,
    color_rgb: Tuple[int,int,int],
    left_punch_margin: float = 0.0,
    border_all_pt: float = 0.0,
) -> Tuple[Path, List[BatesRecord], int, List[Tuple[str,bytes]]]:
    """
    Walk a staged folder tree, apply Bates labels to PDFs and images,
    and return:
      - list of BatesRecord rows (for tables / index)
      - final Bates number used
      - labeled_pairs: (relative_path, bytes) for all labeled files

    NOTE: Output filenames now use the ORIGINAL filename (no 'labeled_' prefix).
    The labeled outputs live under a separate 'labeled' tree so originals are not overwritten.
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir)
        staged = tmp / "staged"
        staged.mkdir(parents=True, exist_ok=True)
        output = tmp / "labeled"
        output.mkdir(parents=True, exist_ok=True)

        # Stage input files in a temp tree
        for disp, data in input_zip_or_pdfs:
            if _is_mac_resource_junk(disp):
                continue
            p = staged / disp
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(data)

        current = start_num
        records: List[BatesRecord] = []
        labeled_pairs: List[Tuple[str, bytes]] = []

        # Walk staged tree and label PDFs + images
        for dirpath, dirnames, filenames in os.walk(staged, topdown=True):
            dirnames[:] = [d for d in sorted(dirnames, key=natural_key) if not _is_mac_resource_junk(d)]
            filenames = [f for f in sorted(filenames, key=natural_key) if not _is_mac_resource_junk(f)]

            rel_dir = str(Path(dirpath).relative_to(staged))
            out_dir = output / rel_dir if rel_dir != "." else output
            out_dir.mkdir(parents=True, exist_ok=True)

            pdfs = [f for f in filenames if f.lower().endswith(".pdf")]
            imgs = [f for f in filenames if Path(f).suffix.lower() in IMAGE_EXTS]

            # --- PDFs ---
            for fname in pdfs:
                src = Path(dirpath) / fname
                # Use ORIGINAL filename in output tree
                out = out_dir / fname
                first = current
                pages_count = 0

                try:
                    reader = PdfReader(str(src))
                    if getattr(reader, "is_encrypted", False):
                        try:
                            reader.decrypt("")
                        except Exception:
                            # Skip encrypted PDFs we can't open
                            continue

                    writer = PdfWriter()
                    for page in reader.pages:
                        w, h = _page_size(page)
                        label = _format_label(prefix, current, digits, with_space=True)
                        overlay = _overlay_pdf(
                            label, w, h, font_name, font_size,
                            margin_right, margin_bottom, color_rgb,
                            left_punch_margin, border_all_pt
                        )
                        page.merge_page(overlay.pages[0])
                        writer.add_page(page)
                        current += 1
                        pages_count += 1

                    with open(out, "wb") as f:
                        writer.write(f)

                    # Add to ZIP pairs using ORIGINAL name/relative path
                    labeled_pairs.append((str(out.relative_to(output)), out.read_bytes()))

                except Exception:
                    # On any error, just skip that file
                    continue

                last = current - 1
                cat = Path(rel_dir).parts[-1] if rel_dir not in (".", "") and Path(rel_dir).parts else ""

                records.append(BatesRecord(
                    rel_dir=rel_dir,
                    filename=fname,  # original filename in Bates table
                    pages_or_files=pages_count,
                    first_label=_format_label(prefix, first, digits, with_space=True),
                    last_label=_format_label(prefix, last, digits, with_space=True),
                    category=cat,
                ))

            # --- Images ---
            for fname in imgs:
                src = Path(dirpath) / fname
                # Use ORIGINAL filename in output tree
                out = out_dir / fname
                first = current

                try:
                    label = _format_label(prefix, current, digits, with_space=True)
                    _label_image(
                        src, out, label, font_name, font_size,
                        margin_right, margin_bottom, color_rgb,
                        left_punch_margin, border_all_pt
                    )
                    labeled_pairs.append((str(out.relative_to(output)), out.read_bytes()))
                    current += 1
                except Exception:
                    continue

                last = current - 1
                cat = Path(rel_dir).parts[-1] if rel_dir not in (".", "") and Path(rel_dir).parts else ""

                records.append(BatesRecord(
                    rel_dir=rel_dir,
                    filename=fname,  # original filename in Bates table
                    pages_or_files=1,
                    first_label=_format_label(prefix, first, digits, with_space=True),
                    last_label=_format_label(prefix, last, digits, with_space=True),
                    category=cat,
                ))

        return records, current - 1, labeled_pairs

def ui_bates():
    st.subheader("Bates Labeler")
    st.caption("Sequential across the entire folder tree.")

    ss = st.session_state
    ss.setdefault("mr_live", 18.0)
    ss.setdefault("mb_live", 18.0)
    ss.setdefault("left_margin_pt", 0.0)
    ss.setdefault("zone_choice", "Bottom Right (Z3)")
    ss.setdefault("zone_padding", 18.0)
    ss.setdefault("border_all_pt", 18.0)

    try:
        poppler_path = _detect_poppler_path()
    except NameError:
        poppler_path = None

    ctrl, view = st.columns([0.95, 1.45], gap="large")

    with ctrl:
        st.markdown("#### Source")
        source = st.radio("Choose source", ["Use current project", "Upload"], horizontal=True, key="bates_source")

        inputs: List[Tuple[str, bytes]] = []
        filenames: List[str] = []

        if source == "Upload":
            mode = st.radio("Input", ["Multiple files", "ZIP of a folder"], horizontal=True, key="bates_mode")
            if mode == "Multiple files":
                files = st.file_uploader("Upload PDFs or images", type=["pdf","jpg","jpeg","png"], accept_multiple_files=True, key="bates_files")
                if files:
                    for f in files:
                        if _is_mac_resource_junk(f.name):
                            continue
                        inputs.append((f.name, f.read()))
                        filenames.append(f.name)
            else:
                z = st.file_uploader("Upload ZIP", type=["zip"], key="bates_zip")
                if z:
                    with zipfile.ZipFile(io.BytesIO(z.read()), "r") as zf:
                        members = [m for m in zf.namelist()
                                   if not m.endswith("/")
                                   and not _is_mac_resource_junk(m)
                                   and any(m.lower().endswith(ext) for ext in [".pdf",".jpg",".jpeg",".png"])]
                        for m in sorted(members, key=natural_key):
                            norm = m.replace("\\", "/")
                            inputs.append((norm, zf.read(m)))
                            filenames.append(norm)
        else:
            if ss.get("project_files"):
                inputs = _filter_pairs_nonjunk(ss["project_files"])
                filenames = [name for name, _ in inputs]
            else:
                st.info("No project files yet. Upload in another tab or set result as current project.")
                inputs = []

        st.divider()
        st.markdown("#### Label")
        c1, c2 = st.columns(2)
        with c1:
            prefix = st.text_input("Prefix", value="J.DOE", key="b_prefix")
            digits = st.number_input("Digits", min_value=3, max_value=12, value=8, step=1, key="b_digits")
            start  = st.number_input("Starting number", min_value=1, value=1, step=1, key="b_start")
        with c2:
            font_size = st.number_input("Font size", min_value=6, max_value=48, value=12, step=1, key="b_font_size")
            font_name = st.text_input("Font name", value="Helvetica", key="b_font_name")
        color_hex = st.color_picker("Color", value="#0000FF", key="b_color")
        color_rgb = _color_from_hex(color_hex)

        st.divider()
        st.markdown("#### Placement")
        zone_choice = st.selectbox(
            "Zone",
            ["Bottom Left (Z1)", "Bottom Center (Z2)", "Bottom Right (Z3)"],
            index=["Bottom Left (Z1)", "Bottom Center (Z2)", "Bottom Right (Z3)"].index(ss.get("zone_choice", "Bottom Right (Z3)")),
            key="b_zone"
        )
        ss["zone_choice"] = zone_choice
        pad_pt = st.number_input("Padding (pt)", min_value=6.0, max_value=144.0, value=float(ss.get("zone_padding", 18.0)), step=1.0, key="b_pad")
        ss["zone_padding"] = float(pad_pt)

        st.divider()
        st.markdown("#### Page Options")
        left_margin_toggle = st.checkbox("Add left margin for 3-hole punch", value=ss["left_margin_pt"] > 0, key="b_left_toggle")
        if left_margin_toggle:
            lm_val = st.number_input("Left margin width (pt)", min_value=6.0, max_value=144.0,
                                     value=float(ss["left_margin_pt"] or 36.0), step=1.0, key="b_left_val")
        else:
            lm_val = 0.0
        ss["left_margin_pt"] = float(lm_val)

        border_toggle = st.checkbox("Add all-sides safety border", value=ss["border_all_pt"] > 0, key="b_border_toggle")
        if border_toggle:
            border_all_pt = st.number_input("Border width (pt)", min_value=6.0, max_value=144.0,
                                            value=float(ss["border_all_pt"] or 18.0), step=1.0, key="b_border_val")
        else:
            border_all_pt = 0.0
        ss["border_all_pt"] = float(border_all_pt)

        try:
            label_sample = _format_label(prefix, int(start), int(digits), with_space=True)
        except NameError:
            label_sample = f"{prefix} {int(start):0{int(digits)}d}"

    with view:
        st.markdown("#### Preview")
        if not inputs:
            st.info("Select a source on the left to enable preview and labeling.")
            return

        sample_name = st.selectbox("Choose a file", filenames, index=0, key="preview_picker")
        sample_bytes = dict(inputs)[sample_name]

        is_pdf = sample_name.lower().endswith(".pdf")
        try:
            poppler_path = _detect_poppler_path()
        except NameError:
            poppler_path = None

        if is_pdf:
            page_count = 1
            try:
                page_count = len(PdfReader(io.BytesIO(sample_bytes)).pages)
            except Exception:
                pass
            page_idx = st.number_input("Page number", min_value=1, max_value=max(1, page_count), value=1, step=1, key="preview_page") - 1
            try:
                img = _preview_image_from_pdf_bytes(sample_bytes, page_idx, poppler_path)
            except TypeError:
                img = _preview_image_from_pdf_bytes(sample_bytes, page_idx)
            if img is None:
                st.info("PDF preview failed. Ensure pdf2image is installed and Poppler is available. Try: pip install pdf2image • brew install poppler.")
                return
        else:
            img = Image.open(io.BytesIO(sample_bytes))
            img = ImageOps.exif_transpose(img)

        dpi = img.info.get("dpi", (150,150))[0] or 150
        px_per_pt = dpi / 72.0

        lm_pt = float(ss.get("left_margin_pt", 0.0))
        if lm_pt > 0:
            lp_px = int(round(lm_pt * px_per_pt))
            canvas = Image.new("RGB", (img.width + lp_px, img.height), (255,255,255))
            canvas.paste(img, (lp_px, 0))
            img = canvas

        border_pt = float(ss.get("border_all_pt", 0.0))
        if border_pt > 0:
            bpx = int(round(border_pt * px_per_pt))
            draw = ImageDraw.Draw(img)
            draw.rectangle([(0, 0), (img.width, bpx)], fill=(255, 255, 255))
            draw.rectangle([(0, img.height - bpx), (img.width, img.height)], fill=(255, 255, 255))
            draw.rectangle([(0, 0), (bpx, img.height)], fill=(255, 255, 255))
            draw.rectangle([(img.width - bpx, 0), (img.width, img.height)], fill=(255, 255, 255))

        def _measure_text_px(sample_img: Image.Image, txt: str, fs: int) -> Tuple[int,int]:
            try:
                font = load_font(font_name, max(8, int(round(fs * px_per_pt))))
            except Exception:
                font = ImageFont.load_default()
            draw = ImageDraw.Draw(sample_img)
            bbox = draw.textbbox((0,0), txt, font=font)
            tw = bbox[2]-bbox[0]; th = bbox[3]-bbox[1]
            return tw, th

        tw, th = _measure_text_px(img, label_sample, int(font_size))

        def compute_margins_from_zone(zone: str, W: int, H: int, TW: int, TH: int, pad_points: float) -> Tuple[float, float]:
            pad_px = pad_points * px_per_pt
            border_px = float(ss.get("border_all_pt", 0.0)) * px_per_pt
            if zone.startswith("Bottom Left"):
                mr_px = max(W - pad_px - TW, border_px)
                mb_px = max(pad_px, border_px)
            elif zone.startswith("Bottom Center"):
                mr_px = max((W - TW) / 2.0, border_px)
                mb_px = max(pad_px, border_px)
            else:
                mr_px = max(pad_px, border_px)
                mb_px = max(pad_px, border_px)
            return mr_px / px_per_pt, mb_px / px_per_pt

        mr_pt, mb_pt = compute_margins_from_zone(ss["zone_choice"], img.width, img.height, tw, th, float(ss["zone_padding"]))
        ss["mr_live"] = float(round(mr_pt, 2))
        ss["mb_live"] = float(round(mb_pt, 2))

        def draw_preview_at(base_img: Image.Image, mx_pt: float, my_pt: float) -> Image.Image:
            im = base_img.copy()
            draw = ImageDraw.Draw(im)
            fs_px = max(8, int(round(int(font_size) * px_per_pt)))
            try:
                font = load_font(font_name, fs_px)
            except Exception:
                font = ImageFont.load_default()
            mx = int(round(mx_pt * px_per_pt))
            my = int(round(my_pt * px_per_pt))
            bbox = draw.textbbox((0,0), label_sample, font=font)
            t_w = bbox[2]-bbox[0]; t_h = bbox[3]-bbox[1]
            x = max(0, im.width - mx - t_w)
            y = max(0, im.height - my - t_h)
            for ox, oy in ((-1,0),(1,0),(0,-1),(0,1)):
                draw.text((x+ox, y+oy), label_sample, font=font, fill=(0,0,0))
            draw.text((x, y), label_sample, font=font, fill=color_rgb)
            return im

        preview_img = draw_preview_at(img, ss["mr_live"], ss["mb_live"])
        st.image(preview_img, caption=f"Preview • {sample_name}", use_container_width=True)

        st.markdown(
            f"**Placement** {ss['zone_choice']} • "
            f"Padding {int(ss['zone_padding'])} pt • "
            f"Border {int(ss['border_all_pt'])} pt • "
            f"Right margin {ss['mr_live']} pt • Bottom margin {ss['mb_live']} pt"
        )

        run_clicked = st.button("Run Bates Labeler", type="primary", use_container_width=True, key="b_run")
        adopt = st.checkbox("Set labeled output as current project for Index", value=True, key="b_adopt")

        if run_clicked:
            records, last_used, labeled_pairs = walk_and_label(
                inputs,
                prefix=prefix,
                start_num=int(start),
                digits=int(digits),
                font_name=font_name,
                font_size=int(font_size),
                margin_right=float(ss["mr_live"]),
                margin_bottom=float(ss["mb_live"]),
                color_rgb=color_rgb,
                left_punch_margin=float(ss["left_margin_pt"]),
                border_all_pt=float(ss["border_all_pt"]),
            )
            zbytes = _zip_from_pairs(labeled_pairs)
            st.success(f"Complete • Final number used: {last_used}")
            _download_bytes("Download labeled files (ZIP)", zbytes, "bates_labeled.zip", "application/zip")

            if records:
                df = pd.DataFrame([r.__dict__ for r in records])
                st.dataframe(df, use_container_width=True)
                ss["bates_records_df"] = df
                ss["labeled_files"] = labeled_pairs
                ss["labeled_zip"] = zbytes
                if adopt and labeled_pairs:
                    project_set(labeled_pairs, origin="Bates Labeler")
                    st.success("Labeled files set as current project.")
            else:
                st.info("No records to display.")

# ---------------- Excel builder ----------------
def build_discovery_xlsx(
    df: pd.DataFrame,
    *,
    party: str = "Client",          # "Client" or "OP"
    title_text: str = "CLIENT NAME - DOCUMENTS",
    date_col_name: str = "Date Produced",
    name_col_name: str = "Document Name/Title",
    cat_col_name: str = "Category",
    bates_col_name: str = "Bates Range",
) -> bytes:
    PARTY_COLORS = {
        "Client": "FFB7DEE8",
        "OP":     "FFFCE4D6",
    }
    category_fill = PatternFill("solid", fgColor=PARTY_COLORS.get(party, "FFB7DEE8"))

    header_fill = PatternFill("solid", fgColor="FF1F4E79")
    header_font = Font(bold=True, color="FFFFFFFF")
    normal_font = Font(bold=False, color="FF000000")
    title_font  = Font(bold=True, size=16)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="FFBBBBBB")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Client" if party == "Client" else "OP"

    ws.merge_cells("A1:C1")
    ws["A1"] = title_text
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    headers = [date_col_name, "Category & Documents provided", "Bated labels"]
    row_idx = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border_thin

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 30

    ws.freeze_panes = "A4"

    for need in [name_col_name, cat_col_name]:
        if need not in df.columns:
            df[need] = ""
    if bates_col_name not in df.columns:
        df[bates_col_name] = ""
    if date_col_name not in df.columns:
        df[date_col_name] = datetime.today().date()

    sdf = df.copy()
    if cat_col_name in sdf.columns:
        sdf[cat_col_name] = sdf[cat_col_name].fillna("")
        sdf.sort_values([cat_col_name, name_col_name], inplace=True, kind="stable")

    row = 4
    for cat, block in sdf.groupby(cat_col_name, dropna=False):
        cat_text = str(cat) if str(cat).strip() else "Uncategorized"
        ws.cell(row=row, column=2, value=cat_text).fill = category_fill
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = left
        for c in (1, 2, 3):
            cc = ws.cell(row=row, column=c); cc.border = border_thin
        row += 1

        for _, r in block.iterrows():
            ws.cell(row=row, column=1, value=r.get(date_col_name, "")).alignment = center
            ws.cell(row=row, column=2, value=r.get(name_col_name, "")).alignment = left
            ws.cell(row=row, column=3, value=r.get(bates_col_name, "")).alignment = left
            for c in (1, 2, 3):
                ws.cell(row=row, column=c).font = normal_font
                ws.cell(row=row, column=c).border = border_thin
            row += 1

        row += 1

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ======================================================
# 4) Discovery Index
# ======================================================
def ui_index():
    st.subheader("Discovery Index")
    st.caption("Create an Excel matching the Master Discovery Spreadsheet style (title, headers, colored category rows).")

    left, right = st.columns([0.95, 1.45], gap="large")

    with left:
        st.markdown("#### Source")
        choices = ["Use last Bates run", "Use current project (labeled)", "Upload labeled ZIP"]
        mode = st.radio("Choose source", choices, horizontal=False, key="idx_source")

        df: Optional[pd.DataFrame] = None
        pairs: Optional[List[Tuple[str, bytes]]] = None

        if mode == "Use last Bates run":
            df = st.session_state.get("bates_records_df")
            if df is None or (isinstance(df, pd.DataFrame) and df.empty):
                st.info("No Bates data in memory yet.")
        elif mode == "Use current project (labeled)":
            pairs = st.session_state.get("labeled_files") or st.session_state.get("project_files")
            pairs = _filter_pairs_nonjunk(pairs or [])
            if not pairs:
                st.info("No labeled project files available yet.")
            else:
                rows = []
                for rel, _b in pairs:
                    p = Path(rel)
                    rel_dir = str(p.parent) if str(p.parent) != "." else ""
                    cat = p.parts[-2] if len(p.parts) > 1 else ""
                    rows.append({"rel_dir": rel_dir, "filename": p.name, "category": cat})
                df = pd.DataFrame(rows)

                det = _scan_pairs_for_bates(pairs)
                if not det.empty:
                    df = df.merge(det[["rel_dir","filename","first_label","last_label"]], on=["rel_dir","filename"], how="left")
        else:
            z = st.file_uploader("Upload labeled ZIP", type=["zip"], key="idx_zip")
            if z:
                pairs = []
                with zipfile.ZipFile(io.BytesIO(z.read()), "r") as zf:
                    rows = []
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        if _is_mac_resource_junk(info.filename):
                            continue
                        rel = info.filename
                        data = zf.read(info)
                        pairs.append((rel, data))
                        p = Path(rel)
                        rel_dir = str(p.parent) if str(p.parent) != "." else ""
                        cat = p.parts[-2] if len(p.parts) > 1 else ""
                        rows.append({"rel_dir": rel_dir, "filename": p.name, "category": cat})
                    df = pd.DataFrame(rows)
                st.success("ZIP analyzed.")
                det = _scan_pairs_for_bates(pairs)
                if not det.empty:
                    df = df.merge(det[["rel_dir","filename","first_label","last_label"]], on=["rel_dir","filename"], how="left")

        st.divider()
        st.markdown("#### Formatting")
        party = st.radio("Color scheme", ["Client", "OP"], horizontal=True, key="idx_party")
        title_text = st.text_input("Sheet title", value="CLIENT NAME - DOCUMENTS", key="idx_title")

        st.divider()
        run = st.button(
            "Build & Export Excel",
            type="primary",
            use_container_width=True,
            disabled=df is None or df.empty,
            key="idx_build_export",
        )

    with right:
        st.markdown("#### Preview / Download")

        if df is None or df.empty:
            if st.session_state.get("bates_records_df") is not None:
                st.info("Select a source and click Build & Export Excel.")
            else:
                st.info("Run the Bates step, select current project (labeled), or upload a labeled ZIP.")
            return

        preview_df = df.copy()

        # Build Bates Range
        if {"first_label","last_label"}.issubset(preview_df.columns):
            fl = preview_df["first_label"].fillna("").astype(str)
            ll = preview_df["last_label"].fillna("").astype(str)
            preview_df["Bates Range"] = np.where(
                (fl != "") & (ll != "") & (fl != ll),
                fl + " - " + ll,
                np.where(fl != "", fl, ll)
            )
        elif "Bates Range" not in preview_df.columns:
            preview_df["Bates Range"] = ""

        # Normalize category/name
        if "category" in preview_df.columns and "filename" in preview_df.columns:
            preview_df.rename(columns={
                "category":"Category",
                "filename":"Document Name/Title"
            }, inplace=True)

        # Date Produced from subfolder
        if "Date Produced" not in preview_df.columns:
            preview_df["Date Produced"] = preview_df.apply(
                lambda r: _extract_date_produced_from_rel(
                    r.get("rel_dir",""),
                    r.get("Document Name/Title","")
                ),
                axis=1
            )
            preview_df["Date Produced"] = preview_df["Date Produced"].apply(
                lambda d: d if pd.notnull(d) and d != "" else datetime.today().date()
            )

        show_cols = ["Date Produced", "Category", "Document Name/Title", "Bates Range"]
        exist_cols = [c for c in show_cols if c in preview_df.columns]
        st.dataframe(preview_df[exist_cols], use_container_width=True, hide_index=True)

        if run:
            try:
                xlsx_bytes = build_discovery_xlsx(
                    preview_df[["Date Produced","Document Name/Title","Category","Bates Range"]],
                    party=party,
                    title_text=title_text,
                )
                st.success("Excel created.")
                st.download_button(
                    "Download discovery.xlsx",
                    data=xlsx_bytes,
                    file_name="discovery.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_discovery_{datetime.now().strftime('%H%M%S')}",
                )
            except Exception as e:
                st.error(f"Failed to build Excel: {e}")

# ======================================================
# 5) REDACTION — engine + hardening patches + SSN context guard
# ======================================================

# --- SSN context: only redact when nearby text hints it's an SSN ---
SSN_CONTEXT_WORDS = re.compile(r"\b(ssn|social\s*security|soc\s*sec|ss#|tin|taxpayer\s*id)\b", re.I)
DEFAULT_REQUIRE_SSN_CONTEXT = True

# Presets updated to support pipes/spaces/hyphens and plain 9-digit SSN
# Note: Removed 9\d\d exclusion from SSN patterns - SSA now assigns 9xx prefixes
PRESETS: Dict[str, List[str]] = {
    "SSN": [
        r"(?<!\d)(?!000|666)\d{3}[-\s|](?!00)\d{2}[-\s|](?!0000)\d{4}(?!\d)",
        r"(?<!\d)(?!000|666)\d{3}(?:(?:\s*\|\s*)|(?:\s+)|(?:-))(?!00)\d{2}(?:(?:\s*\|\s*)|(?:\s+)|(?:-))(?!0000)\d{4}(?!\d)",
        r"(?<!\d)(?!000|666)\d{9}(?!\d)",
    ],
    "Email": [
        r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    ],
    "Phone": [
        r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}",
        r"\+1[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}",
        r"1-\d{3}-\d{3}-\d{4}",
    ],
    "Date": [r"\b(?:\d{1,2}[/-]){2}\d{2,4}\b", r"\b\d{4}-\d{2}-\d{2}\b"],
    "8-digit number": [r"\b\d{8}\b"],
}
ALLOWED_EXTS = {".pdf", ".jpg", ".jpeg", ".png"}

@dataclass
class Hit:
    rel_path: str
    page_num: int
    pattern: str
    matched_text: str

def load_patterns_from_ui(preset_keys: List[str], text_block: str, literals_block: str, case_sensitive: bool) -> List[re.Pattern]:
    raw: List[str] = []
    for key in preset_keys:
        raw.extend(PRESETS.get(key, []))
    if text_block:
        for line in text_block.splitlines():
            s = line.strip()
            if not s or s.startswith('#'):
                continue
            raw.append(s)
    if literals_block:
        for token in re.split(r"[\n,]", literals_block):
            s = token.strip()
            if s:
                raw.append(re.escape(s))
    flags = 0 if case_sensitive else re.IGNORECASE
    compiled = [re.compile(p, flags) for p in raw]
    if not compiled:
        raise ValueError("Provide at least one pattern: select a preset, add regex, or include literal strings.")
    return compiled

def _iter_zip(file: zipfile.ZipFile, allowed_exts: Set[str]) -> Iterable[Tuple[str, bytes]]:
    for info in file.infolist():
        if info.is_dir():
            continue
        if _is_mac_resource_junk(info.filename):
            continue
        ext = Path(info.filename).suffix.lower()
        if ext in allowed_exts:
            yield info.filename, file.read(info)

def image_bytes_to_pdf(img_bytes: bytes) -> bytes:
    with Image.open(io.BytesIO(img_bytes)) as im:
        rgb = im.convert("RGB")
        buf = io.BytesIO()
        rgb.save(buf, format="PDF")
        return buf.getvalue()

PAD_VALUE = 3.0

def _black_fill():
    try:
        from pymupdf import utils as _u  # type: ignore
        return _u.getColor("black")
    except Exception:
        return (0, 0, 0)

def add_black_redaction(page: "fitz.Page", rect: "fitz.Rect", pad: Optional[float] = None) -> None:
    p = float(PAD_VALUE if pad is None else pad)
    r = fitz.Rect(rect)
    r = fitz.Rect(r.x0 - p, r.y0 - p, r.x1 + p, r.y1 + p)
    page.add_redact_annot(r, fill=_black_fill())

def add_black_redaction_leftmask(page: "fitz.Page", rect: "fitz.Rect", pad: Optional[float] = None) -> None:
    p = float(PAD_VALUE if pad is None else pad)
    r = fitz.Rect(rect)
    r = fitz.Rect(r.x0 - p, r.y0 - p, r.x1, r.y1 + p)
    page.add_redact_annot(r, fill=_black_fill())

def prefix_excluding_last_n_digits(s: str, n: int) -> str:
    if n <= 0:
        return s
    digits_seen = 0
    for i in range(len(s) - 1, -1, -1):
        if s[i].isdigit():
            digits_seen += 1
            if digits_seen == n:
                return s[:i]
    return ""

def _repair_pdf_if_needed(raw: bytes) -> bytes:
    if not PIKEPDF_AVAILABLE:
        return raw
    try:
        with pikepdf.open(io.BytesIO(raw)) as pdf:
            buf = io.BytesIO()
            pdf.save(buf)
            return buf.getvalue()
    except Exception:
        return raw

_HYPHENS = ["-", "\u2010", "\u2011", "\u2012", "\u2013", "\u2212"]
_SPACES  = [" ", "\u00A0"]
_PIPES   = ["|", "\u00A6"]

def _search_variants(s: str) -> List[str]:
    s = s.replace("\u200B", "").replace("\u2009", "")
    variants = {s}
    if "-" in s:
        for h in _HYPHENS:
            variants.add(s.replace("-", h))
    for sp in _SPACES:
        variants.add(s.replace(sp, " "))
        variants.add(s.replace(sp, ""))   # compact
    if "|" in s:
        for p in _PIPES:
            variants.add(s.replace("|", p))
        variants.add(s.replace("|", " "))
        variants.add(s.replace("|", ""))
    # (No digits-only variant; prevents false positives)
    return sorted(variants, key=len, reverse=True)

# -------- Core redaction with SSN context option --------
def redact_pdf_bytes(pdf_bytes: bytes, patterns: List[re.Pattern], keep_last_digits: int = 0, *,
                     require_ssn_context: bool = DEFAULT_REQUIRE_SSN_CONTEXT) -> Tuple[bytes, List[Hit]]:
    if fitz is None:
        raise RuntimeError("PyMuPDF (pymupdf) is required. Install with: pip install pymupdf")
    hits: List[Hit] = []

    SSN_PATTERNS = {p for p in PRESETS["SSN"]}
    def _is_ssn_pat(pat: re.Pattern) -> bool:
        return pat.pattern in SSN_PATTERNS

    def _passes_ssn_context_text(full_text: str, m: re.Match) -> bool:
        window = full_text[max(0, m.start()-60): m.end()+60]
        return bool(SSN_CONTEXT_WORDS.search(window))

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception:
        repaired = _repair_pdf_if_needed(pdf_bytes)
        doc = fitz.open(stream=repaired, filetype="pdf")

    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        page_text = page.get_text("text") or ""
        page_had_text = bool(page_text.strip())

        if not page_had_text and pytesseract is not None:
            try:
                pix = page.get_pixmap()
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                ocr = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
                words = ocr.get("text", [])

                # per-word regex matching
                for idx, word in enumerate(words):
                    if not word:
                        continue
                    l = ocr["left"][idx]; t = ocr["top"][idx]
                    w = ocr["width"][idx]; h = ocr["height"][idx]
                    for pat in patterns:
                        if pat.fullmatch(word):
                            # If SSN, require nearby SSN keywords
                            if require_ssn_context and _is_ssn_pat(pat):
                                lo = max(0, idx-6); hi = min(len(words), idx+7)
                                snippet = " ".join(wd for wd in words[lo:hi] if wd)
                                if not SSN_CONTEXT_WORDS.search(snippet or ""):
                                    continue
                            if keep_last_digits > 0:
                                num_digits = sum(ch.isdigit() for ch in word)
                                if num_digits > keep_last_digits:
                                    redact_ratio = (num_digits - keep_last_digits) / max(num_digits, 1)
                                    rect = fitz.Rect(l, t, l + int(w * redact_ratio), t + h)
                                    add_black_redaction_leftmask(page, rect)
                                    hits.append(Hit("", page_index + 1, pat.pattern, word))
                                    continue
                            rect = fitz.Rect(l, t, l + w, t + h)
                            add_black_redaction(page, rect)
                            hits.append(Hit("", page_index + 1, pat.pattern, word))

                # catch SSN split across tokens e.g. 451 | 63 | 1940
                N = len(words)
                def _bbox(i):
                    return (ocr["left"][i], ocr["top"][i], ocr["width"][i], ocr["height"][i])
                def _nearby_has_ssn_keyword(center_i: int) -> bool:
                    lo = max(0, center_i-6); hi = min(N, center_i+7)
                    snippet = " ".join(w for w in words[lo:hi] if w)
                    return bool(SSN_CONTEXT_WORDS.search(snippet or ""))

                for i in range(0, max(0, N - 4)):
                    w0 = words[i]; w1 = words[i+1]; w2 = words[i+2]; w3 = words[i+3]; w4 = words[i+4]
                    if not (w0 and w2 and w4):
                        continue
                    if re.fullmatch(r"\d{3}", w0) and re.fullmatch(r"\D*", w1 or "") \
                       and re.fullmatch(r"\d{2}", w2) and re.fullmatch(r"\D*", w3 or "") \
                       and re.fullmatch(r"\d{4}", w4):
                        if require_ssn_context and not _nearby_has_ssn_keyword(i+2):
                            continue
                        l0,t0,ww0,hh0 = _bbox(i)
                        l1,t1,ww1,hh1 = _bbox(i+1)
                        l2,t2,ww2,hh2 = _bbox(i+2)
                        l3,t3,ww3,hh3 = _bbox(i+3)
                        l4,t4,ww4,hh4 = _bbox(i+4)
                        x0 = min(l0, l1, l2, l3, l4); y0 = min(t0, t1, t2, t3, t4)
                        x1 = max(l0+ww0, l1+ww1, l2+ww2, l3+ww3, l4+ww4); y1 = max(t0+hh0, t1+hh1, t2+hh2, t3+hh3, t4+hh4)
                        add_black_redaction(page, fitz.Rect(x0, y0, x1, y1))
                        hits.append(Hit("", page_index + 1, "OCR_SSN_SPLIT", f"{w0}-{w2}-{w4}"))

            except Exception:
                pass
        else:
            full_targets: List[str] = []
            partial_prefixes: List[str] = []
            for pat in patterns:
                for m in pat.finditer(page_text):
                    s = m.group(0)
                    if not s.strip():
                        continue
                    # SSN: require context around the match window
                    if require_ssn_context and _is_ssn_pat(pat) and not _passes_ssn_context_text(page_text, m):
                        continue
                    if keep_last_digits > 0:
                        prefix = prefix_excluding_last_n_digits(s, keep_last_digits)
                        if prefix:
                            partial_prefixes.append(prefix)
                            hits.append(Hit("", page_index + 1, pat.pattern, s))
                            continue
                    full_targets.append(s)
                    hits.append(Hit("", page_index + 1, pat.pattern, s))

            for s_lit in set(partial_prefixes):
                # keep candidates modest to avoid huge spans
                if len(re.sub(r"\s+", "", s_lit)) > 20:
                    continue
                for candidate in _search_variants(s_lit):
                    try:
                        rects = page.search_for(candidate, quads=True) or []
                        for q in rects:
                            add_black_redaction_leftmask(page, q.rect)
                        if rects:
                            break
                    except Exception:
                        rects = page.search_for(candidate) or []
                        for r in rects:
                            add_black_redaction_leftmask(page, r)
                        if rects:
                            break

            for s_lit in set(full_targets):
                if len(re.sub(r"\s+", "", s_lit)) > 20:
                    continue
                for candidate in _search_variants(s_lit):
                    try:
                        rects = page.search_for(candidate, quads=True) or []
                        for q in rects:
                            add_black_redaction(page, q.rect)
                        if rects:
                            break
                    except Exception:
                        rects = page.search_for(candidate) or []
                        for r in rects:
                            add_black_redaction(page, r)
                        if rects:
                            break

        try:
            page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        except Exception:
            pass

    try:
        doc.apply_redactions()
    except Exception:
        pass

    out = doc.tobytes()
    doc.close()
    return out, hits

def process_zip_bytes(zip_bytes: bytes, patterns: List[re.Pattern], keep_last_digits: int = 0, *,
                      require_ssn_context: bool = DEFAULT_REQUIRE_SSN_CONTEXT) -> Tuple[bytes, List[Hit], Dict]:
    redacted_files: List[Tuple[str, bytes]] = []
    audit_hits: List[Hit] = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zin:
        for rel_path, data in _iter_zip(zin, {".pdf", ".jpg", ".jpeg", ".png"}):
            ext = Path(rel_path).suffix.lower()
            try:
                if ext == ".pdf":
                    red_pdf, hits = redact_pdf_bytes(data, patterns, keep_last_digits, require_ssn_context=require_ssn_context)
                else:
                    pdf_data = image_bytes_to_pdf(data)
                    red_pdf, hits = redact_pdf_bytes(pdf_data, patterns, keep_last_digits, require_ssn_context=require_ssn_context)

                for h in hits:
                    h.rel_path = rel_path
                audit_hits.extend(hits)

                out_name = str(Path(rel_path).with_suffix(".pdf"))
                redacted_files.append((out_name, red_pdf))
            except Exception as e:
                msg = f"Failed to process {rel_path}: {e}"
                redacted_files.append((f"_errors/{rel_path}.txt".replace('..','.'), msg.encode("utf-8")))

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for arcname, data in redacted_files:
            zout.writestr(arcname, data)
        csv_s = io.StringIO()
        cw = csv.writer(csv_s)
        cw.writerow(["file", "page", "pattern", "match"])
        for h in audit_hits:
            cw.writerow([h.rel_path, h.page_num, h.pattern, h.matched_text])
        zout.writestr("audit.csv", csv_s.getvalue().encode("utf-8"))
        report = {
            "files_processed": len(redacted_files),
            "total_hits": len(audit_hits),
            "patterns": [p.pattern for p in patterns],
            "keep_last_digits": keep_last_digits,
            "require_ssn_context": require_ssn_context,
        }
        zout.writestr("report.json", json.dumps(report, indent=2).encode("utf-8"))

    return out_buf.getvalue(), audit_hits, {
        "files_processed": len(redacted_files),
        "total_hits": len(audit_hits),
        "keep_last_digits": keep_last_digits,
        "require_ssn_context": require_ssn_context,
    }

def _zip_project_files_to_bytes(pairs: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rel, b in _filter_pairs_nonjunk(pairs):
            zf.writestr(rel, b)
    buf.seek(0)
    return buf.read()

def ui_redaction():
    st.subheader("Redaction")
    st.caption("SSN/phone/date presets, regex & literals, keep last N digits, audit, ZIP.")

    panel_left, panel_right = st.columns([0.9, 1.6], gap="large")

    with panel_left:
        # --- Source at the top ---
        st.markdown("#### Source")
        st.caption("Allowed types: PDF, JPG, JPEG, PNG")

        source = st.radio(
            "Choose source",
            ["Use current project (auto-zip)", "Upload ZIP"],
            horizontal=True,
            key="red_source",
        )

        uploaded_zip: Optional[bytes] = None
        if source == "Upload ZIP":
            up = st.file_uploader(
                "Upload ZIP",
                type=["zip"],
                accept_multiple_files=False,
                key="red_zip",
            )
            if up:
                uploaded_zip = up.getvalue()
        else:
            pairs = st.session_state.get("project_files") or []
            if not pairs:
                st.info("No current project files. Upload in the Project card or choose Upload ZIP.")
            else:
                uploaded_zip = _zip_project_files_to_bytes(pairs)

        st.divider()

        # --- Redaction pattern configuration ---
        st.markdown("#### Redaction Patterns")

        presets = st.multiselect(
            "Presets",
            options=list(PRESETS.keys()),
            default=["SSN"],
            key="red_presets",
        )
        case_sensitive = st.toggle(
            "Case sensitive",
            value=False,
            key="red_case",
        )

        st.markdown("**Regex patterns** (one per line; `#` for comments)")
        regex_block = st.text_area(
            " ",
            height=140,
            key="red_regex",
            placeholder=r"e.g.\n# SSN pattern\n(?<!\d)(?!000|666|9\d\d)\d{3}[-\s|](?!00)\d{2}[-\s|](?!0000)\d{4}(?!\d)",
        )

        st.markdown("**Literal strings** (comma or newline separated; matched exactly)")
        literals_block = st.text_area(
            "  ",
            height=100,
            key="red_literals",
            placeholder="e.g. A. ROZADA 00000001, JFM - 001212",
        )

        st.markdown("**How many digits to show at the end?**")
        keep_last_digits = st.slider(
            "Keep last N digits visible (mask the rest)",
            min_value=0,
            max_value=6,
            value=0,
            key="red_keep",
            help="If you are redacting account or ID numbers, choose how many of the last digits to keep visible.",
        )

        require_ssn_context = st.checkbox(
            "Only redact SSNs when nearby text says SSN / Social Security (recommended)",
            value=DEFAULT_REQUIRE_SSN_CONTEXT,
            key="red_ssn_ctx",
        )

        st.divider()
        run = st.button(
            "Run Redaction",
            type="primary",
            use_container_width=True,
            disabled=uploaded_zip is None,
            key="red_run",
        )

    with panel_right:
        st.markdown("#### Result")

        if run and uploaded_zip is not None:
            try:
                patterns = load_patterns_from_ui(
                    presets,
                    regex_block,
                    literals_block,
                    case_sensitive,
                )
            except Exception as e:
                st.error(str(e))
                st.stop()

            with st.status("Processing…", expanded=True) as status:
                status.write("Compiling patterns…")
                status.write(f"Using {len(patterns)} pattern(s); keeping last {keep_last_digits} digit(s).")
                try:
                    out_zip_bytes, hits, summary = process_zip_bytes(
                        uploaded_zip,
                        patterns,
                        keep_last_digits,
                        require_ssn_context=require_ssn_context,
                    )
                    status.update(label="Done", state="complete")
                except Exception as e:
                    status.update(label="Failed", state="error")
                    st.exception(e)
                    st.stop()

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Files processed", f"{summary['files_processed']}")
            c2.metric("Total hits", f"{summary['total_hits']}")
            c3.metric("Patterns", f"{len(patterns)}")
            c4.metric("Kept digits", f"{summary['keep_last_digits']}")

            if hits:
                df = pd.DataFrame(
                    [
                        {
                            "rel_path": h.rel_path,
                            "page_num": h.page_num,
                            "pattern": h.pattern,
                            "matched_text": h.matched_text,
                        }
                        for h in hits
                    ]
                )
                st.dataframe(df, use_container_width=True, hide_index=True)
            else:
                st.warning("No matches found. Check your patterns and try again.")

            _download_bytes(
                "Download Redacted ZIP",
                out_zip_bytes,
                "redacted_output.zip",
                "application/zip",
            )
        else:
            st.info("Select a source, configure patterns, then click Run Redaction.")

# ------------------------
# App Layout with Bento Cards & Micro-Animations
# ------------------------
project_init()

# Global styling (Apple-inspired, light + bento cards + micro animations)
st.markdown(
    """
    <style>
        .stApp {
            background-color: #fbfcff;
            color: #111827;
            font-family: -apple-system, BlinkMacSystemFont, system-ui, "SF Pro Text",
                         "Helvetica Neue", Arial, sans-serif;
        }

        /* Main page padding */
        .block-container {
            padding-top: 1.5rem;
            padding-bottom: 2.5rem;
            max-width: 1200px;
        }

        /* Hero card */
        .hero-card {
            background: radial-gradient(circle at 0% 0%, #ffffff 0, #f4f7ff 100%, #e9f0ff 100%);
            border-radius: 22px;
            padding: 22px 26px 20px 26px;
            border: 1px solid #e5e7eb;
            box-shadow: 0 14px 36px rgba(15, 23, 42, 0.06);
            display: flex;
            flex-direction: column;
            gap: 0.4rem;
            margin-bottom: 1.6rem;
            transition: box-shadow 160ms ease, transform 160ms ease;
        }
        .hero-card:hover {
            box-shadow: 0 18px 48px rgba(15, 23, 42, 0.09);
            transform: translateY(-1px);
        }
        .hero-kicker {
            font-size: 11px;
            text-transform: uppercase;
            letter-spacing: .14em;
            color: #6b7280;
            font-weight: 500;
        }
        .hero-title {
            font-size: 24px;
            font-weight: 650;
            color: #020617;
            margin-top: 0.15rem;
        }
        .hero-subtitle {
            font-size: 14px;
            color: #4b5563;
            max-width: 560px;
        }

        /* Bento cards (containers with border=True) */
        div[data-testid="stContainer"] {
            background-color: #ffffff;
            border-radius: 18px;
            border: 1px solid #e5e7eb;
            padding: 1.05rem 1.25rem;
            box-shadow: 0 0 0 rgba(15, 23, 42, 0.02);
            transition:
                box-shadow 150ms ease,
                transform 150ms ease,
                border-color 150ms ease,
                background-color 150ms ease;
        }
        div[data-testid="stContainer"]:hover {
            box-shadow: 0 14px 32px rgba(15, 23, 42, 0.06);
            transform: translateY(-1px);
            border-color: #d1d5db;
            background-color: #fcfcff;
        }

        /* Titles / subheaders */
        h1, h2, h3, h4, h5, h6 {
            font-weight: 600;
            letter-spacing: -0.01em;
        }
        .stMarkdown p {
            font-size: 14.5px;
        }

        /* Hide Streamlit heading anchor icons (link glyphs on hover) */
        h1 a[href^="#"],
        h2 a[href^="#"],
        h3 a[href^="#"],
        h4 a[href^="#"],
        h5 a[href^="#"],
        h6 a[href^="#"] {
            display: none !important;
        }


        /* Tabs: floating pills, no grey background strip, overflow visible */
.stTabs {
    margin-top: 0.6rem;
    overflow: visible;  /* ensure hover shadows aren't clipped */
}

.stTabs [data-baseweb="tab-list"] {
    gap: 0.4rem;
    border-bottom: none;
    background-color: transparent;
    padding: 0;
    border-radius: 0;
    display: inline-flex;
    overflow: visible;  /* key: allow pill shadows to extend */
}

/* Individual tab pills */
.stTabs [data-baseweb="tab"] {
    font-size: 13px;
    font-weight: 500;
    padding: 0.35rem 0.9rem;
    border-radius: 9999px;
    background: transparent;
    color: #4b5563;
    border: none;
    box-shadow: none;
    transition:
        background-color 140ms ease,
        color 140ms ease,
        transform 120ms ease,
        box-shadow 140ms ease;
}
.stTabs [data-baseweb="tab"]:hover {
    background-color: #dde4ff;
    color: #1f2937;
    transform: translateY(-1px);
    box-shadow: 0 4px 10px rgba(15, 23, 42, 0.15);
}
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    background-color: #007aff;
    color: #ffffff;
    box-shadow: 0 6px 16px rgba(0, 122, 255, 0.45);
    transform: translateY(-1px);
}


        /* Buttons – subtle pill look, no loud accents */
        .stButton > button {
            border-radius: 9999px;
            background-color: #e5e7eb;
            color: #111827;
            border: 1px solid #d1d5db;
            padding: 0.42rem 1.2rem;
            font-size: 13.5px;
            font-weight: 500;
            box-shadow: 0 3px 8px rgba(15, 23, 42, 0.08);
            transition:
                background-color 140ms ease,
                color 140ms ease,
                transform 120ms ease,
                box-shadow 140ms ease,
                border-color 140ms ease;
        }
        .stButton > button:hover {
            background-color: #d4d7dd;
            border-color: #c4c7ce;
            transform: translateY(-1px);
            box-shadow: 0 6px 16px rgba(15, 23, 42, 0.14);
        }
        .stButton > button:active {
            transform: translateY(0);
            box-shadow: 0 2px 6px rgba(15, 23, 42, 0.10);
        }

        /* Metrics subtle styling */
        [data-testid="stMetric"] {
            padding: 0.25rem 0.3rem;
        }

        /* File uploader micro-outline on hover */
        [data-testid="stFileUploader"] > div {
            border-radius: 12px !important;
            border-color: #d1d5db !important;
            transition:
                border-color 150ms ease,
                background-color 150ms ease,
                box-shadow 150ms ease,
                transform 120ms ease;
        }
        [data-testid="stFileUploader"] > div:hover {
            border-color: #007aff !important;
            background-color: #f3f6ff !important;
            box-shadow: 0 6px 20px rgba(15, 23, 42, 0.08);
            transform: translateY(-1px);
        }

        /* Radios / selects feel a bit lighter */
        .stRadio > label, .stSelectbox > label {
            font-size: 13px;
            color: #4b5563;
        }

        /* Status boxes */
        [data-testid="stStatus"] {
            border-radius: 14px !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# Hero / Header
st.markdown(
    """
    <div class="hero-card">
        <div class="hero-kicker">Ramage Law Group · Internal Toolkit</div>
        <div class="hero-title">Discovery One-Stop</div>
    </div>
    """,
    unsafe_allow_html=True,
)

# Single universal Project card:
#   - If no project loaded: show upload controls (ZIP or multiple files)
#   - Once project exists: show project status + simple reset button
with st.container(border=True):
    st.markdown("##### Project")

    if not st.session_state.get("project_files"):
        st.caption("Upload a production set once and use it across all tools.")

        proj_mode = st.radio(
            "Input type",
            ["ZIP of a folder", "Multiple files"],
            horizontal=True,
            key="proj_mode",
        )

        pairs: List[Tuple[str, bytes]] = []

        if proj_mode == "ZIP of a folder":
            proj_zip = st.file_uploader(
                "Upload project ZIP",
                type=["zip"],
                key="proj_zip",
                help="Upload a single ZIP containing your discovery folder tree.",
            )
            if proj_zip:
                with zipfile.ZipFile(io.BytesIO(proj_zip.read()), "r") as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        if _is_mac_resource_junk(info.filename):
                            continue
                        pairs.append((info.filename, zf.read(info)))
        else:
            proj_files = st.file_uploader(
                "Upload PDFs or images",
                type=["pdf", "jpg", "jpeg", "png"],
                accept_multiple_files=True,
                key="proj_files",
                help="You can select multiple files at once.",
            )
            if proj_files:
                for f in proj_files:
                    if _is_mac_resource_junk(f.name):
                        continue
                    pairs.append((f.name, f.read()))

        set_btn = st.button(
            "Set as current project",
            use_container_width=False,
            key="proj_set_btn",
            disabled=not pairs,
        )

        if set_btn and pairs:
            project_set(pairs, origin="Project upload")
            st.success(f"Project loaded with {len(pairs)} file(s).")
    else:
        # Project already exists – show status + simple reset
        st.write(project_summary())
        cols = st.columns([1, 0.25])
        with cols[1]:
            if st.button("Reset project", use_container_width=True, key="reset_btn"):
                project_clear()
                st.success("Project reset. Upload a new set to start fresh.")
        if st.session_state.get("last_action"):
            st.caption(st.session_state.get("last_action"))

# Tabs — Redaction before Bates, pill-style like buttons
tab_unlock, tab_year, tab_redact, tab_bates, tab_index = st.tabs(
    ["Unlock PDFs", "Organize by Year", "Redaction", "Bates Labeler", "Discovery Index"]
)

with tab_unlock:
    with st.container(border=True):
        ui_unlocker()

with tab_year:
    with st.container(border=True):
        ui_organizer()

with tab_redact:
    with st.container(border=True):
        ui_redaction()

with tab_bates:
    with st.container(border=True):
        ui_bates()

with tab_index:
    with st.container(border=True):
        ui_index()

st.markdown("---")
st.markdown(
    "<span style='color:#6b7280;font-size:13px;'>Tip: Use “Set result as current project” to carry files forward across tabs without re-uploading.</span>",
    unsafe_allow_html=True,
)
