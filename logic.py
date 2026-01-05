from __future__ import annotations

import io, os, re, csv, zipfile, tempfile, shutil, json, platform
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Optional, Tuple, List, Iterable, Set
from collections import Counter

import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF / imaging
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.colors import Color
from PIL import Image, ImageDraw, ImageFont, ImageOps

# Optional: pdf2image (for previews & OCR fallback)
try:
    from pdf2image import convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except Exception:
    PDF2IMAGE_AVAILABLE = False

# Optional: pikepdf for unlocking & repair
try:
    import pikepdf
    from pikepdf import PasswordError, PdfError
    PIKEPDF_AVAILABLE = True
except Exception:
    pikepdf = None
    PasswordError = PdfError = Exception
    PIKEPDF_AVAILABLE = False

# Optional: PyMuPDF
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

# Optional: OCR
try:
    import pytesseract
except Exception:
    pytesseract = None

# ------------------------
# Global helpers
# ------------------------
def load_font(font_name: str, size: int) -> ImageFont.FreeTypeFont:
    """
    Load a font from common system paths or fallback to default.
    """
    # Common paths for Arial or similar sans-serif
    candidates = [
        # User requested specific path (if provided as absolute path)
        font_name,
        # macOS
        f"/Library/Fonts/{font_name}.ttf",
        f"/System/Library/Fonts/{font_name}.ttf",
        # Linux (Debian/Ubuntu/Render)
        f"/usr/share/fonts/truetype/msttcorefonts/{font_name}.ttf",
        f"/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", # Fallback 1
        f"/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", # Fallback 2
    ]
    
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
            
    # If all else fails, try loading by name (OS might resolve it)
    try:
        return ImageFont.truetype(f"{font_name}.ttf", size)
    except Exception:
        pass
        
    return ImageFont.load_default()

def _zip_dir(dir_path: Path) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in dir_path.rglob("*"):
            if p.is_file():
                zf.write(p, p.relative_to(dir_path))
    buf.seek(0)
    return buf.read()

def _zip_from_pairs(pairs: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rel, b in pairs:
            if rel.endswith("/"):
                continue
            zf.writestr(rel, b)
    buf.seek(0)
    return buf.read()

def _color_from_hex(hex_str: str) -> Tuple[int, int, int]:
    hex_str = hex_str.strip("#")
    if len(hex_str) == 6:
        r = int(hex_str[0:2], 16)
        g = int(hex_str[2:4], 16)
        b = int(hex_str[4:6], 16)
        return (r, g, b)
    return (0, 0, 255)

def natural_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]

def _detect_poppler_path() -> Optional[str]:
    for c in ("/opt/homebrew/bin", "/usr/local/bin"):
        if Path(c, "pdftoppm").exists():
            return c
    return None

def _format_label(prefix: str, number: int, digits: int, with_space: bool = True) -> str:
    return f"{prefix}{' ' if with_space else ''}{number:0{digits}d}"

def _pil_dpi(image: Image.Image) -> float:
    dpi = image.info.get("dpi")
    if isinstance(dpi, tuple) and dpi and dpi[0]:
        return float(dpi[0])
    return 150.0

# ---------- macOS junk filters ----------
def _is_mac_resource_junk(path_str: str) -> bool:
    p = str(path_str).replace("\\", "/")
    base = Path(p).name
    if "/__MACOSX/" in p:
        return True
    if base.startswith("._"):
        return True
    if base.lower() == ".ds_store":
        return True
    return False

def _filter_pairs_nonjunk(pairs: List[Tuple[str, bytes]]) -> List[Tuple[str, bytes]]:
    return [(rel, b) for rel, b in pairs if not _is_mac_resource_junk(rel)]

# --- Date Produced helper: parse from 2nd-level subfolder under the parent ---
def _parse_date_from_text(text: str) -> Optional[date]:
    if not text:
        return None
    # yyyy.mm.dd | yyyy-mm-dd | yyyy_mm_dd
    for m in re.finditer(r"\b(20\d{2}|19\d{2})[._/-](\d{1,2})[._/-](\d{1,2})\b", text):
        y, mo, d = map(int, m.groups())
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            pass
    # yyyymmdd
    m = re.search(r"\b(20\d{2}|19\d{2})(\d{2})(\d{2})\b", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            pass
    # mm.dd.yyyy | etc
    for m in re.finditer(r"\b(\d{1,2})[._/-](\d{1,2})[._/-](20\d{2}|19\d{2})\b", text):
        mo, d, y = map(int, m.groups())
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            pass
    return None

def _extract_date_produced_from_rel(rel_dir: str, filename: str = "") -> Optional[date]:
    parts = Path(rel_dir).parts if rel_dir else ()
    candidates: List[str] = []
    # Explicit: 2nd-level inside the parent folder (index 1)
    if len(parts) >= 2:
        candidates.append(parts[1])
    # fallbacks
    candidates.extend(parts)
    if filename:
        candidates.append(filename)
    for token in candidates:
        d = _parse_date_from_text(token)
        if d:
            return d
    return None

# ---------------- Smarter Bates detection (avoid random numbers) ----------------
_BLACKLIST_PREFIXES = {
    "MONTHLY", "BOX", "ID", "TARGET", "REQUESTED", "MISC"
}

_CANDIDATE_BATES_RE = re.compile(
    r"\b([A-Z][A-Z0-9. ]{1,30}?)[\s\-–—]*([0-9]{6,10})\b"
)

def _normalize_prefix(s: str) -> str:
    s = re.sub(r"\s+", " ", s).strip(" -–—.")
    return s.upper()

def _is_zero_padded(num: str) -> bool:
    return len(num) >= 6 and num[0] == "0"

def _extract_candidates(text: str) -> List[Tuple[str, str]]:
    out: List[Tuple[str,str]] = []
    if not text:
        return out
    for m in _CANDIDATE_BATES_RE.finditer(text.upper()):
        pfx = _normalize_prefix(m.group(1))
        num = m.group(2)
        out.append((pfx, num))
    return out

def _choose_dominant_prefix(cands: List[Tuple[str, str]]) -> Optional[str]:
    if not cands:
        return None
    zp = [p for p,n in cands if _is_zero_padded(n) and p not in _BLACKLIST_PREFIXES]
    if zp:
        return Counter(zp).most_common(1)[0][0]
    nb = [p for p,_ in cands if p not in _BLACKLIST_PREFIXES]
    if nb:
        return Counter(nb).most_common(1)[0][0]
    return None

def _best_token_for_prefix(cands: List[Tuple[str,str]], want_prefix: str) -> Optional[str]:
    for pfx, num in cands:
        if pfx == want_prefix and _is_zero_padded(num):
            return f"{pfx} {num}"
    for pfx, num in cands:
        if pfx == want_prefix:
            return f"{pfx} {num}"
    return None

def _pdf_page_text_or_ocr(pdf_bytes: bytes, page_index_zero: int) -> str:
    if fitz is not None:
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            if 0 <= page_index_zero < doc.page_count:
                page = doc.load_page(page_index_zero)
                txt = page.get_text("text") or ""
                if txt.strip():
                    doc.close()
                    return txt
                if pytesseract is not None:
                    pix = page.get_pixmap()
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    ocr_txt = pytesseract.image_to_string(img) or ""
                    doc.close()
                    return ocr_txt
            doc.close()
        except Exception:
            pass
    if PDF2IMAGE_AVAILABLE and pytesseract is not None:
        try:
            imgs = convert_from_bytes(pdf_bytes, first_page=page_index_zero+1, last_page=page_index_zero+1, dpi=200)
            if imgs:
                return pytesseract.image_to_string(imgs[0]) or ""
        except Exception:
            pass
    return ""

def _image_bytes_text_ocr(img_bytes: bytes) -> str:
    try:
        with Image.open(io.BytesIO(img_bytes)) as im:
            im = ImageOps.exif_transpose(im)
            if pytesseract is not None:
                return pytesseract.image_to_string(im) or ""
    except Exception:
        pass
    return ""

def _extract_bates_for_file(rel_path: str, data: bytes) -> Tuple[Optional[str], Optional[str]]:
    ext = Path(rel_path).suffix.lower()

    if ext == ".pdf":
        txt1 = _pdf_page_text_or_ocr(data, 0)
        c1 = _extract_candidates(txt1)

        page_count = None
        if fitz is not None:
            try:
                d = fitz.open(stream=data, filetype="pdf"); page_count = d.page_count; d.close()
            except Exception:
                pass
        if page_count is None:
            try:
                page_count = len(PdfReader(io.BytesIO(data)).pages)
            except Exception:
                page_count = 1
        last_index = max(0, page_count - 1)

        txtN = _pdf_page_text_or_ocr(data, last_index)
        cN = _extract_candidates(txtN)

        all_cands = c1 + cN
        dom = _choose_dominant_prefix(all_cands)
        if not dom:
            return None, None

        first_tok = _best_token_for_prefix(c1, dom) or _best_token_for_prefix(all_cands, dom)
        last_tok  = _best_token_for_prefix(cN, dom) or _best_token_for_prefix(all_cands[::-1], dom)

        if not first_tok and not last_tok:
            return None, None
        if first_tok and not last_tok:
            return first_tok, first_tok
        if last_tok and not first_tok:
            return last_tok, last_tok

        def _num_part(tok: str) -> int:
            return int(re.search(r"(\d{6,10})$", tok).group(1))

        try:
            n1 = _num_part(first_tok); n2 = _num_part(last_tok)
            if n2 < n1:
                first_tok, last_tok = last_tok, first_tok
        except Exception:
            pass
        return first_tok, last_tok

    elif ext in {".jpg", ".jpeg", ".png"}:
        txt = _image_bytes_text_ocr(data)
        cands = _extract_candidates(txt)
        dom = _choose_dominant_prefix(cands)
        if dom:
            tok = _best_token_for_prefix(cands, dom)
            return tok, tok
        return None, None

    else:
        return None, None

def scan_pairs_for_bates(pairs: List[Tuple[str, bytes]]) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for i, (rel, b) in enumerate(pairs, start=1):
        if _is_mac_resource_junk(rel):
            continue
        p = Path(rel)
        rel_dir = str(p.parent) if str(p.parent) != "." else ""
        fname = p.name
        try:
            first, last = _extract_bates_for_file(rel, b)
        except Exception:
            first, last = None, None
        rows.append({
            "rel_dir": rel_dir,
            "filename": fname,
            "first_label": first or "",
            "last_label": last or ""
        })
    return pd.DataFrame(rows)

# ======================================================
# 1) Unlock PDFs
# ======================================================
def unlock_pdfs(files: List[Tuple[str, bytes]], password_mode: str, password_for_all: Optional[str], password_map: Dict[str, str]) -> bytes:
    if not PIKEPDF_AVAILABLE:
        raise RuntimeError("pikepdf is not installed")

    def _resolve_password(path: str) -> Optional[str]:
        if password_mode == "Single password for all":
            return (password_for_all or "").strip() or None
        elif password_mode == "Per-file password list (CSV)":
            base = os.path.basename(path)
            stem = os.path.splitext(base)[0]
            return password_map.get(path) or password_map.get(base) or password_map.get(stem)
        else:
            return None

    def _process_pdf(src_bytes: bytes, password: Optional[str]) -> Tuple[str, Optional[bytes]]:
        try:
            with io.BytesIO(src_bytes) as src_buf:
                try:
                    pdf = pikepdf.open(src_buf) if password is None else pikepdf.open(src_buf, password=password)
                except PasswordError:
                    return "Password required or incorrect", None
                except PdfError as e:
                    return f"PDF error: {e.__class__.__name__}", None
                out_mem = io.BytesIO()
                pdf.save(out_mem)  # saved without encryption
                pdf.close()
                return "Unlocked", out_mem.getvalue()
        except Exception as e:
            return f"Unexpected error: {e}", None

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files:
            if _is_mac_resource_junk(fname):
                continue
            if fname.lower().endswith(".pdf"):
                status, unlocked_data = _process_pdf(data, _resolve_password(fname))
                out_name = os.path.splitext(fname)[0] + "_unlocked.pdf"
                if unlocked_data is not None:
                    zf.writestr(out_name, unlocked_data)
            elif fname.lower().endswith(".zip"):
                try:
                    with zipfile.ZipFile(io.BytesIO(data), 'r') as inzip:
                        for member in inzip.namelist():
                            if member.endswith('/'):
                                continue
                            if _is_mac_resource_junk(member):
                                continue
                            if not member.lower().endswith('.pdf'):
                                continue
                            pw = _resolve_password(member)
                            status, unlocked_data = _process_pdf(inzip.read(member), pw)
                            out_name = f"{os.path.splitext(member)[0]}_unlocked.pdf"
                            if unlocked_data is not None:
                                zf.writestr(out_name, unlocked_data)
                except zipfile.BadZipFile:
                    pass
    
    zip_buffer.seek(0)
    return zip_buffer.read()

# ======================================================
# 2) Organize by Year
# ======================================================
MONTHS = r"Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?"
PATTERNS: List[re.Pattern] = [
    re.compile(rf"(?<!\d)((?P<year>20\d{{2}}|19\d{{2}}))(?!\d)", re.IGNORECASE),
    re.compile(rf"(?<!\d)\d{{1,2}}[ _\.-]\d{{1,2}}[ _\.-](?P<year>20\d{{2}}|19\d{{2}})(?!\d)", re.IGNORECASE),
    re.compile(rf"(?:(?:{MONTHS})\s*\d{{1,2}}[ ,._-]*)?(?P<year>20\d{{2}}|19\d{{2}})", re.IGNORECASE),
]

def preprocess_filename(name: str) -> str:
    return re.sub(r"^[A-Za-z]*\d{4,}[ _\.-]*", "", name)

def extract_year_from_name(name: str, min_year: int, max_year: int, year_policy: str = "first") -> Tuple[Optional[int], str]:
    name = preprocess_filename(name)
    candidates: List[Tuple[int, str, Tuple[int, int]]] = []
    for idx, pat in enumerate(PATTERNS, start=1):
        for m in pat.finditer(name):
            y = int(m.group("year"))
            if min_year <= y <= max_year:
                candidates.append((y, f"pattern{idx}", m.span()))
    if not candidates:
        return None, "no-year-found"
    if year_policy == "max":
        chosen = max(candidates, key=lambda t: t[0])
    elif year_policy == "last":
        chosen = candidates[-1]
    else:
        chosen = candidates[0]
    year, patname, span = chosen
    return year, f"{patname}@{span}"

def organize_by_year(files: List[Tuple[str, bytes]], min_year: int, max_year: int, year_policy: str, unknown_folder: str) -> bytes:
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir)
        out_root = tmp / f"organized_{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        out_root.mkdir(parents=True, exist_ok=True)

        for display_name, data in files:
            year, reason = extract_year_from_name(Path(display_name).name, min_year, max_year, year_policy)
            folder = str(year) if year is not None else unknown_folder
            target_dir = out_root / folder
            target_dir.mkdir(parents=True, exist_ok=True)
            dest = target_dir / Path(display_name).name
            i = 1
            while dest.exists():
                dest = target_dir / f"{Path(display_name).stem}__{i}{Path(display_name).suffix}"
                i += 1
            dest.write_bytes(data)

        return _zip_dir(out_root)

# ======================================================
# 3) Bates Labeler
# ======================================================
IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

@dataclass
class BatesRecord:
    rel_dir: str
    filename: str
    pages_or_files: int
    first_label: str
    last_label: str
    category: str  # deepest folder

def _page_size(page) -> Tuple[float, float]:
    return float(page.mediabox.width), float(page.mediabox.height)

def _overlay_pdf(
    label: str,
    w: float, h: float,
    font_name: str, font_size: int,
    margin_right: float, margin_bottom: float,
    color_rgb: Tuple[int,int,int],
    left_punch_margin: float = 0.0,
    border_all_pt: float = 0.0,
) -> PdfReader:
    from io import BytesIO
    r, g, b = color_rgb
    packet = BytesIO()
    can = rl_canvas.Canvas(packet, pagesize=(w, h))

    if border_all_pt and border_all_pt > 0:
        can.setFillColor(Color(1, 1, 1))
        B = float(border_all_pt)
        can.rect(0, h - B, w, B, stroke=0, fill=1)
        can.rect(0, 0, w, B, stroke=0, fill=1)
        can.rect(0, 0, B, h, stroke=0, fill=1)
        can.rect(w - B, 0, B, h, stroke=0, fill=1)

    if left_punch_margin and left_punch_margin > 0:
        can.setFillColor(Color(1, 1, 1))
        can.rect(0, 0, left_punch_margin, h, stroke=0, fill=1)

    can.setFont(font_name, font_size)
    can.setFillColor(Color(r/255, g/255, b/255))
    eff_mr = max(margin_right, border_all_pt or 0.0)
    eff_mb = max(margin_bottom, border_all_pt or 0.0)
    can.drawRightString(w - eff_mr, eff_mb, label)

    can.save()
    packet.seek(0)
    return PdfReader(packet)

def _label_image(
    in_file: Path, out_file: Path, label: str,
    font_name: str, font_size_pt: int,
    margin_right_pt: float, margin_bottom_pt: float,
    color_rgb: Tuple[int,int,int],
    left_punch_margin_pt: float = 0.0,
    border_all_pt: float = 0.0,
):
    img = Image.open(in_file)
    img = ImageOps.exif_transpose(img)

    dpi = _pil_dpi(img)
    px_per_point = dpi / 72.0

    mx = int(round(margin_right_pt * px_per_point))
    my = int(round(margin_bottom_pt * px_per_point))
    lp = int(round(left_punch_margin_pt * px_per_point))
    bp = int(round(border_all_pt * px_per_point))

    if lp > 0:
        new_img = Image.new(
            "RGB" if img.mode == "RGB" else "RGBA",
            (img.width + lp, img.height),
            (255, 255, 255) if img.mode != "RGBA" else (255, 255, 255, 0),
        )
        new_img.paste(img, (lp, 0))
        img = new_img

    if bp > 0:
        new_img = Image.new(
            "RGB" if img.mode == "RGB" else "RGBA",
            (img.width + 2 * bp, img.height + 2 * bp),
            (255, 255, 255) if img.mode != "RGBA" else (255, 255, 255, 0),
        )
        new_img.paste(img, (bp, bp))
        img = new_img
        mx = max(mx, bp)
        my = max(my, bp)

    fs_from_points = font_size_pt * px_per_point
    relative_min = 0.025 * min(img.width, img.height)
    fs_px = int(max(10, round(max(fs_from_points, relative_min))))

    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGBA")

    draw = ImageDraw.Draw(img)
    try:
        font = load_font(font_name, fs_px)
    except Exception:
        font = ImageFont.load_default()

    bbox = draw.textbbox((0, 0), label, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    x = max(0, img.width - mx - tw)
    y = max(0, img.height - my - th)

    for ox, oy in ((-1, 0), (1, 0), (0, -1), (0, 1)):
        draw.text((x + ox, y + oy), label, font=font, fill=(0, 0, 0))

    draw.text((x, y), label, font=font, fill=color_rgb)

    if out_file.suffix.lower() in [".jpg", ".jpeg"]:
        img.convert("RGB").save(out_file, quality=92, optimize=True)
    else:
        img.save(out_file)

def _measure_text_px(txt: str, font_name: str, font_size_px: int) -> Tuple[int, int]:
    try:
        font = load_font(font_name, font_size_px)
    except Exception:
        font = ImageFont.load_default()
    # Create a dummy image to draw on
    dummy = Image.new("RGB", (10, 10))
    draw = ImageDraw.Draw(dummy)
    bbox = draw.textbbox((0, 0), txt, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    return tw, th

def _compute_margins_for_page(
    zone: str, w: float, h: float,
    text: str, font_name: str, font_size: int,
    padding_pt: float, border_pt: float
) -> Tuple[float, float]:
    # Convert points to pixels (assuming 72 DPI for PDF coordinate system match)
    # Actually, ReportLab uses points directly (1/72 inch).
    # PIL uses pixels. If we treat PDF points as "pixels" for PIL measurement:
    
    # Measure text in "points" (treating them as pixels for the font loader)
    # This is an approximation, but consistent with how _overlay_pdf works
    tw, th = _measure_text_px(text, font_name, font_size)
    
    pad = padding_pt
    border = border_pt
    
    if zone.startswith("Bottom Left"):
        mr = max(w - pad - tw, border)
        mb = max(pad, border)
    elif zone.startswith("Bottom Center"):
        mr = max((w - tw) / 2.0, border)
        mb = max(pad, border)
    else: # Bottom Right (default)
        mr = max(pad, border)
        mb = max(pad, border)
        
    return mr, mb

def walk_and_label(
    input_zip_or_pdfs: List[Tuple[str, bytes]], *,
    prefix: str, start_num: int, digits: int,
    font_name: str, font_size: int,
    margin_right: float = 18.0, margin_bottom: float = 18.0,
    zone: Optional[str] = None, zone_padding: float = 18.0,
    color_rgb: Tuple[int,int,int],
    left_punch_margin: float = 0.0,
    border_all_pt: float = 0.0,
) -> Tuple[List[BatesRecord], int, List[Tuple[str,bytes]]]:
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir)
        staged = tmp / "staged"
        staged.mkdir(parents=True, exist_ok=True)
        output = tmp / "labeled"
        output.mkdir(parents=True, exist_ok=True)

        for disp, data in input_zip_or_pdfs:
            if _is_mac_resource_junk(disp):
                continue
            p = staged / disp
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(data)

        current = start_num
        records: List[BatesRecord] = []
        labeled_pairs: List[Tuple[str, bytes]] = []

        for dirpath, dirnames, filenames in os.walk(staged, topdown=True):
            dirnames[:] = [d for d in sorted(dirnames, key=natural_key) if not _is_mac_resource_junk(d)]
            filenames = [f for f in sorted(filenames, key=natural_key) if not _is_mac_resource_junk(f)]

            rel_dir = str(Path(dirpath).relative_to(staged))
            out_dir = output / rel_dir if rel_dir != "." else output
            out_dir.mkdir(parents=True, exist_ok=True)

            pdfs = [f for f in filenames if f.lower().endswith(".pdf")]
            imgs = [f for f in filenames if Path(f).suffix.lower() in IMAGE_EXTS]

            for fname in pdfs:
                src = Path(dirpath) / fname
                out = out_dir / fname
                first = current
                pages_count = 0

                try:
                    reader = PdfReader(str(src))
                    if getattr(reader, "is_encrypted", False):
                        try:
                            reader.decrypt("")
                        except Exception:
                            continue

                    writer = PdfWriter()
                    for page in reader.pages:
                        w, h = _page_size(page)
                        label = _format_label(prefix, current, digits, with_space=True)
                        
                        # Calculate margins dynamically if zone is provided
                        mr, mb = margin_right, margin_bottom
                        if zone:
                            mr, mb = _compute_margins_for_page(
                                zone, w, h, label, font_name, font_size, zone_padding, border_all_pt
                            )

                        overlay = _overlay_pdf(
                            label, w, h, font_name, font_size,
                            mr, mb, color_rgb,
                            left_punch_margin, border_all_pt
                        )
                        page.merge_page(overlay.pages[0])
                        writer.add_page(page)
                        current += 1
                        pages_count += 1

                    with open(out, "wb") as f:
                        writer.write(f)

                    labeled_pairs.append((str(out.relative_to(output)), out.read_bytes()))

                except Exception:
                    continue

                last = current - 1
                cat = Path(rel_dir).parts[-1] if rel_dir not in (".", "") and Path(rel_dir).parts else ""

                records.append(BatesRecord(
                    rel_dir=rel_dir,
                    filename=fname,
                    pages_or_files=pages_count,
                    first_label=_format_label(prefix, first, digits, with_space=True),
                    last_label=_format_label(prefix, last, digits, with_space=True),
                    category=cat,
                ))

            for fname in imgs:
                src = Path(dirpath) / fname
                out = out_dir / fname
                first = current

                try:
                    label = _format_label(prefix, current, digits, with_space=True)
                    
                    # For images, we need to open it to get dimensions for zone calculation
                    # _label_image opens it again, but that's okay for now.
                    # Or we can let _label_image handle it if we passed zone?
                    # For now, let's just stick to the passed fixed margins for images 
                    # OR we can open it here quickly.
                    
                    mr, mb = margin_right, margin_bottom
                    if zone:
                        with Image.open(io.BytesIO(src.read_bytes())) as tmp_img:
                            tmp_img = ImageOps.exif_transpose(tmp_img)
                            # DPI adjustment logic from _label_image is complex to replicate here perfectly
                            # without refactoring _label_image.
                            # However, _label_image takes POINTS for margins.
                            # Our _compute_margins_for_page returns POINTS.
                            # So we just need width/height in POINTS? 
                            # No, _compute_margins_for_page assumes W/H are comparable to font size units.
                            
                            # Let's simplify: For images, _label_image does its own drawing.
                            # It doesn't support "Zone" param natively yet.
                            # But we can calculate the margins if we know the image size.
                            
                            # Actually, _label_image converts points to pixels using DPI.
                            # If we pass margins in points, it handles it.
                            # So we just need to know the image width/height in POINTS to center it?
                            # Image width in points = width_px / (dpi/72).
                            
                            dpi = _pil_dpi(tmp_img)
                            px_per_pt = dpi / 72.0
                            w_pt = tmp_img.width / px_per_pt
                            h_pt = tmp_img.height / px_per_pt
                            
                            mr, mb = _compute_margins_for_page(
                                zone, w_pt, h_pt, label, font_name, font_size, zone_padding, border_all_pt
                            )

                    _label_image(
                        src, out, label, font_name, font_size,
                        mr, mb, color_rgb,
                        left_punch_margin, border_all_pt
                    )
                    labeled_pairs.append((str(out.relative_to(output)), out.read_bytes()))
                    current += 1
                except Exception:
                    continue

                last = current - 1
                cat = Path(rel_dir).parts[-1] if rel_dir not in (".", "") and Path(rel_dir).parts else ""

                records.append(BatesRecord(
                    rel_dir=rel_dir,
                    filename=fname,
                    pages_or_files=1,
                    first_label=_format_label(prefix, first, digits, with_space=True),
                    last_label=_format_label(prefix, last, digits, with_space=True),
                    category=cat,
                ))

    return records, current - 1, labeled_pairs

# ---------------- Excel builder ----------------
def build_discovery_xlsx(
    df: pd.DataFrame,
    *,
    party: str = "Client",
    title_text: str = "CLIENT NAME - DOCUMENTS",
    date_col_name: str = "Date Produced",
    name_col_name: str = "Document Name/Title",
    cat_col_name: str = "Category",
    bates_col_name: str = "Bates Range",
) -> bytes:
    PARTY_COLORS = {
        "Client": "FFB7DEE8",
        "OP":     "FFFCE4D6",
    }
    category_fill = PatternFill("solid", fgColor=PARTY_COLORS.get(party, "FFB7DEE8"))

    header_fill = PatternFill("solid", fgColor="FF1F4E79")
    header_font = Font(bold=True, color="FFFFFFFF")
    normal_font = Font(bold=False, color="FF000000")
    title_font  = Font(bold=True, size=16)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="FFBBBBBB")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Client" if party == "Client" else "OP"

    ws.merge_cells("A1:C1")
    ws["A1"] = title_text
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    headers = [date_col_name, "Category & Documents provided", "Bated labels"]
    row_idx = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border_thin

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 30

    ws.freeze_panes = "A4"

    for need in [name_col_name, cat_col_name]:
        if need not in df.columns:
            df[need] = ""
    if bates_col_name not in df.columns:
        df[bates_col_name] = ""
    if date_col_name not in df.columns:
        df[date_col_name] = datetime.today().date()

    sdf = df.copy()
    if cat_col_name in sdf.columns:
        sdf[cat_col_name] = sdf[cat_col_name].fillna("")
        sdf.sort_values([cat_col_name, name_col_name], inplace=True, kind="stable")

    row = 4
    for cat, block in sdf.groupby(cat_col_name, dropna=False):
        cat_text = str(cat) if str(cat).strip() else "Uncategorized"
        ws.cell(row=row, column=2, value=cat_text).fill = category_fill
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = left
        for c in (1, 2, 3):
            cc = ws.cell(row=row, column=c); cc.border = border_thin
        row += 1

        for _, r in block.iterrows():
            ws.cell(row=row, column=1, value=r.get(date_col_name, "")).alignment = center
            ws.cell(row=row, column=2, value=r.get(name_col_name, "")).alignment = left
            ws.cell(row=row, column=3, value=r.get(bates_col_name, "")).alignment = left
            for c in (1, 2, 3):
                ws.cell(row=row, column=c).font = normal_font
                ws.cell(row=row, column=c).border = border_thin
            row += 1

        row += 1

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ======================================================
# 5) REDACTION
# ======================================================
SSN_CONTEXT_WORDS = re.compile(r"\b(ssn|social\s*security|soc\s*sec|ss#|tin|taxpayer\s*id)\b", re.I)
DEFAULT_REQUIRE_SSN_CONTEXT = True

# Presets updated to support pipes/spaces/hyphens and plain 9-digit SSN
# Note: Removed 9\d\d exclusion from SSN patterns - SSA now assigns 9xx prefixes
PRESETS: Dict[str, List[str]] = {
    "SSN": [
        r"(?<!\d)(?!000|666)\d{3}[-\s|](?!00)\d{2}[-\s|](?!0000)\d{4}(?!\d)",
        r"(?<!\d)(?!000|666)\d{3}(?:(?:\s*\|\s*)|(?:\s+)|(?:-))(?!00)\d{2}(?:(?:\s*\|\s*)|(?:\s+)|(?:-))(?!0000)\d{4}(?!\d)",
        r"(?<!\d)(?!000|666)\d{9}(?!\d)",
    ],
    "Email": [
        r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    ],
    "Phone": [
        r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}",
        r"\+1[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}",
        r"1-\d{3}-\d{3}-\d{4}",
    ],
    "Date": [r"\b(?:\d{1,2}[/-]){2}\d{2,4}\b", r"\b\d{4}-\d{2}-\d{2}\b"],
    "8-digit number": [r"\b\d{8}\b"],
}

@dataclass
class Hit:
    rel_path: str
    page_num: int
    pattern: str
    matched_text: str

def load_patterns(preset_keys: List[str], text_block: str, literals_block: str, case_sensitive: bool) -> List[re.Pattern]:
    raw: List[str] = []
    for key in preset_keys:
        raw.extend(PRESETS.get(key, []))
    if text_block:
        for line in text_block.splitlines():
            s = line.strip()
            if not s or s.startswith('#'):
                continue
            raw.append(s)
    if literals_block:
        for token in re.split(r"[\n,]", literals_block):
            s = token.strip()
            if s:
                raw.append(re.escape(s))
    flags = 0 if case_sensitive else re.IGNORECASE
    compiled = [re.compile(p, flags) for p in raw]
    if not compiled:
        raise ValueError("Provide at least one pattern: select a preset, add regex, or include literal strings.")
    return compiled

def _iter_zip(file: zipfile.ZipFile, allowed_exts: Set[str]) -> Iterable[Tuple[str, bytes]]:
    for info in file.infolist():
        if info.is_dir():
            continue
        if _is_mac_resource_junk(info.filename):
            continue
        ext = Path(info.filename).suffix.lower()
        if ext in allowed_exts:
            yield info.filename, file.read(info)

def image_bytes_to_pdf(img_bytes: bytes) -> bytes:
    with Image.open(io.BytesIO(img_bytes)) as im:
        rgb = im.convert("RGB")
        buf = io.BytesIO()
        rgb.save(buf, format="PDF")
        return buf.getvalue()

PAD_VALUE = 3.0

def _black_fill():
    try:
        from pymupdf import utils as _u
        return _u.getColor("black")
    except Exception:
        return (0, 0, 0)

def add_black_redaction(page: "fitz.Page", rect: "fitz.Rect", pad: Optional[float] = None) -> None:
    p = float(PAD_VALUE if pad is None else pad)
    r = fitz.Rect(rect)
    r = fitz.Rect(r.x0 - p, r.y0 - p, r.x1 + p, r.y1 + p)
    page.add_redact_annot(r, fill=_black_fill())

def add_black_redaction_leftmask(page: "fitz.Page", rect: "fitz.Rect", pad: Optional[float] = None) -> None:
    p = float(PAD_VALUE if pad is None else pad)
    r = fitz.Rect(rect)
    r = fitz.Rect(r.x0 - p, r.y0 - p, r.x1, r.y1 + p)
    page.add_redact_annot(r, fill=_black_fill())

def prefix_excluding_last_n_digits(s: str, n: int) -> str:
    if n <= 0:
        return s
    digits_seen = 0
    for i in range(len(s) - 1, -1, -1):
        if s[i].isdigit():
            digits_seen += 1
            if digits_seen == n:
                return s[:i]
    return ""

def _repair_pdf_if_needed(raw: bytes) -> bytes:
    if not PIKEPDF_AVAILABLE:
        return raw
    try:
        with pikepdf.open(io.BytesIO(raw)) as pdf:
            buf = io.BytesIO()
            pdf.save(buf)
            return buf.getvalue()
    except Exception:
        return raw

_HYPHENS = ["-", "\u2010", "\u2011", "\u2012", "\u2013", "\u2212"]
_SPACES  = [" ", "\u00A0"]
_PIPES   = ["|", "\u00A6"]

def _search_variants(s: str) -> List[str]:
    s = s.replace("\u200B", "").replace("\u2009", "")
    variants = {s}
    if "-" in s:
        for h in _HYPHENS:
            variants.add(s.replace("-", h))
    for sp in _SPACES:
        variants.add(s.replace(sp, " "))
        variants.add(s.replace(sp, ""))
    if "|" in s:
        for p in _PIPES:
            variants.add(s.replace("|", p))
        variants.add(s.replace("|", " "))
        variants.add(s.replace("|", ""))
    return sorted(variants, key=len, reverse=True)

def redact_pdf_bytes(pdf_bytes: bytes, patterns: List[re.Pattern], keep_last_digits: int = 0, *,
                     require_ssn_context: bool = DEFAULT_REQUIRE_SSN_CONTEXT) -> Tuple[bytes, List[Hit]]:
    if fitz is None:
        raise RuntimeError("PyMuPDF (pymupdf) is required. Install with: pip install pymupdf")
    hits: List[Hit] = []

    SSN_PATTERNS = {p for p in PRESETS["SSN"]}
    def _is_ssn_pat(pat: re.Pattern) -> bool:
        return pat.pattern in SSN_PATTERNS

    def _passes_ssn_context_text(full_text: str, m: re.Match) -> bool:
        window = full_text[max(0, m.start()-60): m.end()+60]
        return bool(SSN_CONTEXT_WORDS.search(window))

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception:
        repaired = _repair_pdf_if_needed(pdf_bytes)
        doc = fitz.open(stream=repaired, filetype="pdf")

    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        page_text = page.get_text("text") or ""
        page_had_text = bool(page_text.strip())

        if not page_had_text and pytesseract is not None:
            try:
                pix = page.get_pixmap()
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                ocr = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
                words = ocr.get("text", [])

                for idx, word in enumerate(words):
                    if not word:
                        continue
                    l = ocr["left"][idx]; t = ocr["top"][idx]
                    w = ocr["width"][idx]; h = ocr["height"][idx]
                    for pat in patterns:
                        if pat.fullmatch(word):
                            if require_ssn_context and _is_ssn_pat(pat):
                                lo = max(0, idx-6); hi = min(len(words), idx+7)
                                snippet = " ".join(wd for wd in words[lo:hi] if wd)
                                if not SSN_CONTEXT_WORDS.search(snippet or ""):
                                    continue
                            if keep_last_digits > 0:
                                num_digits = sum(ch.isdigit() for ch in word)
                                if num_digits > keep_last_digits:
                                    redact_ratio = (num_digits - keep_last_digits) / max(num_digits, 1)
                                    rect = fitz.Rect(l, t, l + int(w * redact_ratio), t + h)
                                    add_black_redaction_leftmask(page, rect)
                                    hits.append(Hit("", page_index + 1, pat.pattern, word))
                                    continue
                            rect = fitz.Rect(l, t, l + w, t + h)
                            add_black_redaction(page, rect)
                            hits.append(Hit("", page_index + 1, pat.pattern, word))

                N = len(words)
                def _bbox(i):
                    return (ocr["left"][i], ocr["top"][i], ocr["width"][i], ocr["height"][i])
                def _nearby_has_ssn_keyword(center_i: int) -> bool:
                    lo = max(0, center_i-6); hi = min(N, center_i+7)
                    snippet = " ".join(w for w in words[lo:hi] if w)
                    return bool(SSN_CONTEXT_WORDS.search(snippet or ""))

                for i in range(0, max(0, N - 4)):
                    w0 = words[i]; w1 = words[i+1]; w2 = words[i+2]; w3 = words[i+3]; w4 = words[i+4]
                    if not (w0 and w2 and w4):
                        continue
                    if re.fullmatch(r"\d{3}", w0) and re.fullmatch(r"\D*", w1 or "") \
                       and re.fullmatch(r"\d{2}", w2) and re.fullmatch(r"\D*", w3 or "") \
                       and re.fullmatch(r"\d{4}", w4):
                        if require_ssn_context and not _nearby_has_ssn_keyword(i+2):
                            continue
                        l0,t0,ww0,hh0 = _bbox(i)
                        l1,t1,ww1,hh1 = _bbox(i+1)
                        l2,t2,ww2,hh2 = _bbox(i+2)
                        l3,t3,ww3,hh3 = _bbox(i+3)
                        l4,t4,ww4,hh4 = _bbox(i+4)
                        x0 = min(l0, l1, l2, l3, l4); y0 = min(t0, t1, t2, t3, t4)
                        x1 = max(l0+ww0, l1+ww1, l2+ww2, l3+ww3, l4+ww4); y1 = max(t0+hh0, t1+hh1, t2+hh2, t3+hh3, t4+hh4)
                        add_black_redaction(page, fitz.Rect(x0, y0, x1, y1))
                        hits.append(Hit("", page_index + 1, "OCR_SSN_SPLIT", f"{w0}-{w2}-{w4}"))

            except Exception:
                pass
        else:
            full_targets: List[str] = []
            partial_prefixes: List[str] = []
            for pat in patterns:
                for m in pat.finditer(page_text):
                    s = m.group(0)
                    if not s.strip():
                        continue
                    if require_ssn_context and _is_ssn_pat(pat) and not _passes_ssn_context_text(page_text, m):
                        continue
                    if keep_last_digits > 0:
                        prefix = prefix_excluding_last_n_digits(s, keep_last_digits)
                        if prefix:
                            partial_prefixes.append(prefix)
                            hits.append(Hit("", page_index + 1, pat.pattern, s))
                            continue
                    full_targets.append(s)
                    hits.append(Hit("", page_index + 1, pat.pattern, s))

            for s_lit in set(partial_prefixes):
                # increased from 20 to 60 for emails
                if len(re.sub(r"\s+", "", s_lit)) > 60:
                    continue
                for candidate in _search_variants(s_lit):
                    try:
                        rects = page.search_for(candidate, quads=True) or []
                        for q in rects:
                            add_black_redaction_leftmask(page, q.rect)
                        if rects:
                            break
                    except Exception:
                        rects = page.search_for(candidate) or []
                        for r in rects:
                            add_black_redaction_leftmask(page, r)
                        if rects:
                            break

            for s_lit in set(full_targets):
                # increased from 20 to 60 for emails
                if len(re.sub(r"\s+", "", s_lit)) > 60:
                    continue
                for candidate in _search_variants(s_lit):
                    try:
                        rects = page.search_for(candidate, quads=True) or []
                        for q in rects:
                            add_black_redaction(page, q.rect)
                        if rects:
                            break
                    except Exception:
                        rects = page.search_for(candidate) or []
                        for r in rects:
                            add_black_redaction(page, r)
                        if rects:
                            break

        try:
            page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        except Exception:
            pass

    try:
        doc.apply_redactions()
    except Exception:
        pass

    out = doc.tobytes()
    doc.close()
    return out, hits

def process_zip_bytes(zip_bytes: bytes, patterns: List[re.Pattern], keep_last_digits: int = 0, *,
                      require_ssn_context: bool = DEFAULT_REQUIRE_SSN_CONTEXT) -> Tuple[bytes, List[Hit], Dict]:
    redacted_files: List[Tuple[str, bytes]] = []
    audit_hits: List[Hit] = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zin:
        for rel_path, data in _iter_zip(zin, {".pdf", ".jpg", ".jpeg", ".png"}):
            ext = Path(rel_path).suffix.lower()
            try:
                if ext == ".pdf":
                    red_pdf, hits = redact_pdf_bytes(data, patterns, keep_last_digits, require_ssn_context=require_ssn_context)
                else:
                    pdf_data = image_bytes_to_pdf(data)
                    red_pdf, hits = redact_pdf_bytes(pdf_data, patterns, keep_last_digits, require_ssn_context=require_ssn_context)

                for h in hits:
                    h.rel_path = rel_path
                audit_hits.extend(hits)

                out_name = str(Path(rel_path).with_suffix(".pdf"))
                redacted_files.append((out_name, red_pdf))
            except Exception as e:
                msg = f"Failed to process {rel_path}: {e}"
                redacted_files.append((f"_errors/{rel_path}.txt".replace('..','.'), msg.encode("utf-8")))

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for arcname, data in redacted_files:
            zout.writestr(arcname, data)
        csv_s = io.StringIO()
        cw = csv.writer(csv_s)
        cw.writerow(["file", "page", "pattern", "match"])
        for h in audit_hits:
            cw.writerow([h.rel_path, h.page_num, h.pattern, h.matched_text])
        zout.writestr("audit.csv", csv_s.getvalue().encode("utf-8"))
        report = {
            "files_processed": len(redacted_files),
            "total_hits": len(audit_hits),
            "patterns": [p.pattern for p in patterns],
            "keep_last_digits": keep_last_digits,
            "require_ssn_context": require_ssn_context,
        }
        zout.writestr("report.json", json.dumps(report, indent=2).encode("utf-8"))

    return out_buf.getvalue(), audit_hits, {
        "files_processed": len(redacted_files),
        "total_hits": len(audit_hits),
        "keep_last_digits": keep_last_digits,
        "require_ssn_context": require_ssn_context,
    }
